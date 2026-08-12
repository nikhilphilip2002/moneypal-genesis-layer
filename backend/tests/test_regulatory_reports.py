"""Integrity tests for the five-output RBI report registry and clean templates."""

import io
import re

import openpyxl
import pytest

from app.services import regulatory_reports as reports


PAN_RE = re.compile(r"\b[A-Z]{5}\d{4}[A-Z]\b")
EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.]+")


def _sample_data(report_id: str):
    definition = reports.REPORTS[report_id]
    return {
        "report_id": report_id,
        "return_code": definition.return_code,
        "name": definition.name,
        "frequency": definition.frequency,
        "start_date": "2026-04-01" if definition.frequency == "quarterly" else "2026-06-01",
        "end_date": "2026-06-30",
        "cashflows_thousands": {
            "P": [1.0] * 10,
            "I": [0.5] * 10,
            "total": [1.5] * 10,
        },
    }


def test_registry_exposes_five_report_outputs():
    assert list(reports.REPORTS) == [
        "dnbs02", "dnbs13", "dnbs4a", "dnbs4b_structural", "dnbs4b_irs"
    ]


def test_new_clean_templates_preserve_expected_sheets():
    expected = {
        "DNBS13_Blank_Template.xlsx": ["FilingInfo", "DNBS13", "AuthorisedSignatory"],
        "DNBS4A_Blank_Template.xlsx": [
            "FilingInfo", "DNBS4AShortTermDynamicLiquidity", "AuthorisedSignatory"
        ],
        "DNBS4B_Blank_Template.xlsx": [
            "Filing Info", "AuthorisedSignatory", "DNBS4BStructuralLiquidity", "DNBS4BIRS"
        ],
    }
    for filename, sheets in expected.items():
        wb = openpyxl.load_workbook(reports.ASSET_DIR / filename, data_only=False)
        assert wb.sheetnames == sheets


def test_new_templates_contain_no_pan_or_email():
    for filename in (
        "DNBS13_Blank_Template.xlsx", "DNBS4A_Blank_Template.xlsx", "DNBS4B_Blank_Template.xlsx"
    ):
        wb = openpyxl.load_workbook(reports.ASSET_DIR / filename, data_only=False)
        values = "\n".join(
            str(cell.value)
            for ws in wb.worksheets for row in ws.iter_rows() for cell in row
            if cell.value is not None
        )
        assert not PAN_RE.search(values)
        assert not EMAIL_RE.search(values)


def test_dnbs4a_writer_discovers_rows_and_bucket_headers():
    definition = reports.REPORTS["dnbs4a"]
    wb = openpyxl.load_workbook(reports.ASSET_DIR / definition.template)
    reports._write_dnbs4a(wb, definition, _sample_data("dnbs4a"))
    ws = wb["DNBS4AShortTermDynamicLiquidity"]
    row = reports._find_row(ws, "6 Interest inflow on performing Advances", (38, 88))
    assert [ws.cell(row, col).value for col in range(3, 8)] == [0.5, 0.5, 0.5, 1.0, 0.5]
    assert ws.cell(row, 8).value == 3.0


def test_dnbs4b_writer_populates_both_statements_in_the_original_workbook():
    definition = reports.REPORTS["dnbs4b_structural"]
    wb = openpyxl.load_workbook(reports.ASSET_DIR / definition.template)
    reports._write_dnbs4b(wb, definition, _sample_data("dnbs4b_structural"))
    assert wb["DNBS4BStructuralLiquidity"]["M155"].value == 15.0
    assert wb["DNBS4BIRS"]["P164"].value == 15.0


def test_dnbs13_writer_keeps_detail_rows_blank():
    definition = reports.REPORTS["dnbs13"]
    wb = openpyxl.load_workbook(reports.ASSET_DIR / definition.template)
    data = _sample_data("dnbs13")
    reports._write_dnbs13(wb, definition, data)
    output = io.BytesIO()
    wb.save(output)
    reopened = openpyxl.load_workbook(io.BytesIO(output.getvalue()))
    assert all(reopened["DNBS13"].cell(13, col).value is None for col in range(2, 18))


def test_custom_range_requires_both_dates():
    definition = reports.REPORTS["dnbs4b_structural"]
    with pytest.raises(reports.ReportError, match="both start_date and end_date"):
        reports._parse_request(definition, "custom", "", "2026-07-01", None)


def test_custom_range_rejects_reversed_dates():
    definition = reports.REPORTS["dnbs4b_structural"]
    with pytest.raises(reports.ReportError, match="precedes period start"):
        reports._parse_request(
            definition, "custom", "", "2026-07-31", "2026-07-01"
        )


def test_custom_range_accepts_iso_dates_without_changing_report_frequency():
    definition = reports.REPORTS["dnbs4a"]
    assert reports._parse_request(
        definition, "custom", "", "2026-07-01", "2026-07-31"
    ) == ("2026-07-01", "2026-07-31")


def test_custom_report_is_explicitly_not_filing_eligible(monkeypatch):
    monkeypatch.setattr(
        reports,
        "_build_alm_report",
        lambda definition, start, end: {
            "status": "complete",
            "start_date": start,
            "end_date": end,
        },
    )
    data = reports.get_report_data(
        "dnbs4b_structural",
        "custom",
        "",
        "2026-07-01",
        "2026-07-31",
    )
    assert data["report_mode"] == "custom"
    assert data["filing_eligible"] is False
    assert "internal analytical" in data["filing_note"]


def test_custom_workbook_is_stamped_as_internal():
    definition = reports.REPORTS["dnbs13"]
    wb = openpyxl.load_workbook(reports.ASSET_DIR / definition.template)
    reports._fill_common_workbook_metadata(
        wb, definition, "2026-07-01", "2026-07-31", "custom"
    )
    values = {
        str(cell.value)
        for row in wb["FilingInfo"].iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert "Custom (Internal)" in values
