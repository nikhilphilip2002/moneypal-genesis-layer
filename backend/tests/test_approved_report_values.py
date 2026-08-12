from pathlib import Path

import openpyxl
import pytest

from app.services.approved_report_values import (
    ApprovedReportValueError,
    apply_approved_values,
)


def _workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "Template heading"
    ws["C3"].number_format = "0.00"
    return wb


def test_approved_values_fill_blank_cells_and_preserve_style():
    wb = _workbook()
    style_id = wb["Report"]["C3"].style_id
    count = apply_approved_values(
        wb,
        [{"sheet_name": "Report", "target_cell": "C3", "value": 12.5}],
    )
    assert count == 1
    assert wb["Report"]["C3"].value == 12.5
    assert wb["Report"]["C3"].style_id == style_id


def test_approved_values_cannot_overwrite_template_or_calculated_values():
    wb = _workbook()
    with pytest.raises(ApprovedReportValueError, match="cannot overwrite"):
        apply_approved_values(
            wb,
            [{"sheet_name": "Report", "target_cell": "A1", "value": "replacement"}],
        )


def test_approved_values_reject_unknown_sheets():
    with pytest.raises(ApprovedReportValueError, match="unknown sheet"):
        apply_approved_values(
            _workbook(),
            [{"sheet_name": "Missing", "target_cell": "A1", "value": 1}],
        )


def test_reference_migration_requires_one_value_and_maker_checker_separation():
    sql = Path("backend/scripts/sql/regulatory_reference_tables.sql").read_text()
    assert "num_nonnulls(value_numeric, value_text, value_date, value_boolean) = 1" in sql
    assert "checker <> maker" in sql
    assert "approved_at timestamp with time zone NOT NULL" in sql
