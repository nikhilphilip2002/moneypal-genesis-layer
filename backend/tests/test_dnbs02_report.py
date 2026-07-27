"""Tests for the RBI DNBS-02 return generator.

The tests that matter here are the ones the previous suite lacked: that figures land on
the RBI line they claim to be, that a section with no source stays blank instead of being
invented, and that the template carries no prior filing's data.
"""

import io
import re

import openpyxl
import pytest

from app.services import dnbs02_service as svc
from app.services.dnbs02_service import (
    CellMapError,
    PeriodError,
    _find_label_row,
    _norm,
    generate_dnbs02_excel,
    get_dnbs02_report_data,
    get_template_path,
    parse_period_range,
    validate_cell_map,
    TABLE_BLOCKS,
)

PAN_RE = re.compile(r"\b[A-Z]{5}\d{4}[A-Z]\b")
EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.]+")

# Names, PANs and figures that appear only in the filed return the old template was built
# from. None of them may ever reach a generated workbook.
PRIOR_FILING_MARKERS = [
    "SUBRAMANYA",
    "MEGHARAJ",
    "GEO ENGINEERING",
    "RHINESTONE",
    "CCI LTD",
    "CHITRAKALA",
    "giccltd",
    "AL CARGO",
    "FRNKLIN",
    "HINDUJA GLOBAL",
    "PROSPER FINANCIAL",
    "Virtual Branch",
    "District Desk",
]


def _db_available() -> bool:
    try:
        with svc._db_cursor() as (_conn, cur):
            cur.execute("SELECT 1")
            return True
    except Exception:
        return False


requires_db = pytest.mark.skipif(not _db_available(), reason="PostgreSQL warehouse not reachable")


class TestPeriodParsing:
    def test_monthly(self):
        assert parse_period_range("monthly", "2026-05") == ("2026-05-01", "2026-05-31")

    def test_monthly_leap_february(self):
        assert parse_period_range("monthly", "2028-02") == ("2028-02-01", "2028-02-29")

    def test_quarterly_is_fiscal(self):
        # Q1 of FY2026 is Apr-Jun 2026; Q4 spills into the next calendar year.
        assert parse_period_range("quarterly", "2026-Q1") == ("2026-04-01", "2026-06-30")
        assert parse_period_range("quarterly", "2025-Q4") == ("2026-01-01", "2026-03-31")

    def test_yearly(self):
        assert parse_period_range("yearly", "2025-2026") == ("2025-04-01", "2026-03-31")
        assert parse_period_range("yearly", "2025") == ("2025-04-01", "2026-03-31")

    @pytest.mark.parametrize(
        "frequency,period",
        [
            ("monthly", "junk"),
            ("monthly", "2026-13"),
            ("monthly", ""),
            ("quarterly", "2025-Q9"),
            ("quarterly", "nonsense"),
            ("yearly", "2025-2030"),
            ("weekly", "2026-05"),
        ],
    )
    def test_malformed_periods_raise(self, frequency, period):
        """A bad period must fail, not silently become a hardcoded default."""
        with pytest.raises(PeriodError):
            parse_period_range(frequency, period)


class TestTemplateIntegrity:
    def test_template_is_the_blank(self):
        assert get_template_path().endswith(svc.TEMPLATE_FILENAME)

    def test_template_carries_no_prior_filing_data(self):
        """The shipped template must not be a completed return."""
        wb = openpyxl.load_workbook(get_template_path())
        offences = []
        for name in wb.sheetnames:
            for row in wb[name].iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value)
                    if PAN_RE.search(text):
                        offences.append(f"{name}!{cell.coordinate} PAN {text!r}")
                    if EMAIL_RE.search(text):
                        offences.append(f"{name}!{cell.coordinate} email {text!r}")
                    if isinstance(cell.value, (int, float)) and cell.row >= 13:
                        offences.append(f"{name}!{cell.coordinate} figure {cell.value!r}")
        assert not offences, "prior filing data in template: " + "; ".join(offences[:10])

    def test_second_tier_column_headers_survive_blanking(self):
        """Blanking must strip reported figures without taking header labels with them:
        PART8A's Min / Max / Weighted Average and PART8C's Amount sit in row 13."""
        wb = openpyxl.load_workbook(get_template_path())
        part8a = wb["DNBS02_PART8A"]
        assert _norm(part8a["G13"].value) == "min"
        assert _norm(part8a["H13"].value) == "max"
        assert _norm(part8a["I13"].value).startswith("weighted average")
        assert all(_norm(wb["DNBS02_PART8C"][f"{c}13"].value) == "amount" for c in "CD")

    def test_cell_map_matches_template(self):
        """Every declared column must sit under the header it claims."""
        validate_cell_map(openpyxl.load_workbook(get_template_path()))

    def test_cell_map_detects_a_shifted_column(self):
        """Guard the guard: a moved header must be caught, not written through."""
        wb = openpyxl.load_workbook(get_template_path())
        block = TABLE_BLOCKS[0]
        wb[block.sheet][f"{block.columns[0].column}{block.header_row}"] = "Something Else"
        with pytest.raises(CellMapError):
            validate_cell_map(wb)

    def test_label_lookup_rejects_ambiguity(self):
        wb = openpyxl.load_workbook(get_template_path())
        sheet = wb["DNBS02_PART3"]
        # "(a) Interest" prefixes both "(a) Interest on Inter-corporate Deposits" and
        # the investment-income "(a) Interest", so it must be rejected unless scoped.
        with pytest.raises(CellMapError):
            _find_label_row(sheet, "(a) Interest")
        assert _find_label_row(sheet, "(a) Interest", within=svc.PART3_INVESTMENT_SCOPE)

    def test_label_lookup_rejects_missing_label(self):
        wb = openpyxl.load_workbook(get_template_path())
        with pytest.raises(CellMapError):
            _find_label_row(wb["DNBS02_PART1"], "No Such RBI Line")


class TestAssetClassification:
    def test_sma_codes_are_not_npas(self):
        """SMA-0/1/2 are standard assets under stress; treating SMA-2 as a loss asset
        was inflating both NPAs and provisions."""
        for code in ("STD", "SMA0", "SMA1", "SMA2"):
            assert code not in svc.NPA_ASSET_CODES
            assert code in svc.STANDARD_ASSET_CODES

    def test_npa_codes_are_classified(self):
        for code in ("SUB", "DBT", "LOSS"):
            assert code in svc.NPA_ASSET_CODES

    def test_every_known_code_has_a_label(self):
        for code in svc.NPA_ASSET_CODES + svc.STANDARD_ASSET_CODES:
            assert code in svc.ASSET_CODE_LABELS


@pytest.fixture(scope="module")
def report():
    return get_dnbs02_report_data(frequency="monthly", period="2026-05")


@pytest.fixture(scope="module")
def workbook():
    return openpyxl.load_workbook(
        io.BytesIO(generate_dnbs02_excel(frequency="monthly", period="2026-05"))
    )


@requires_db
class TestReportData:
    def test_reports_on_the_requested_period_end(self, report):
        assert report["end_date"] == "2026-05-31"
        assert report["snapshot_date"] == "2026-05-31"

    def test_unbacked_period_raises(self):
        """A period with no snapshot must error rather than produce numbers."""
        with pytest.raises(PeriodError):
            get_dnbs02_report_data(frequency="monthly", period="2019-01")

    def test_every_section_declares_provenance(self, report):
        assert report["provenance"]
        for name, entry in report["provenance"].items():
            assert entry["status"] in {"ok", "empty", "error", "no_source"}, name
            if entry["status"] in {"error", "no_source"}:
                assert entry.get("error"), f"{name} must explain why it has no data"

    def test_is_live_only_when_all_sections_resolved(self, report):
        expected = bool(report["live_sections"]) and not report["degraded_sections"]
        assert report["is_live_pg"] is expected

    def test_sections_without_a_source_are_empty(self, report):
        """No fabricated stand-ins for sections the warehouse cannot back."""
        assert report["provenance"]["annex2_shareholders"]["status"] == "no_source"
        assert report["annex2_shareholders"] == []
        assert report["provenance"]["annex10_investment_entities"]["status"] == "no_source"
        for inv in report["annex10_top_investments"]:
            assert inv["entity_name"] == ""
            assert inv["pan"] == ""

    def test_crar_is_null_without_risk_weighted_assets(self, report):
        assert report["summary"]["crar_pct"] is None

    def test_no_npas_are_reported(self, report):
        """This portfolio holds only STD/SMA assets, so GNPA must be zero."""
        assert report["summary"]["gross_npa_amount"] == 0.0
        assert report["summary"]["gross_npa_pct"] == 0.0
        assert report["annex11_top_npas"] == []
        assert all(not row["is_npa"] for row in report["part8_asset_quality"])

    def test_annex9_is_aggregated_by_borrower(self, report):
        borrowers = report["annex9_top_borrowers"]
        assert 0 < len(borrowers) <= 25
        ids = [b["cust_id"] for b in borrowers]
        assert len(ids) == len(set(ids)), "a borrower must appear at most once"
        assert borrowers == sorted(borrowers, key=lambda b: b["total_outstanding"], reverse=True)

    def test_annex9_carries_real_pans(self, report):
        pans = [b["pan"] for b in report["annex9_top_borrowers"]]
        assert all(PAN_RE.fullmatch(p) for p in pans if p), "PANs must be real, not 'NA'"

    def test_part8_amounts_reconcile_with_the_loan_book(self, report):
        total = sum(row["amount_lakhs"] for row in report["part8_asset_quality"])
        assert total == pytest.approx(report["summary"]["total_loan_book"], abs=0.05)

    def test_msme_uses_the_loan_master_rate_column(self, report):
        """Part 8A is sourced from genlnacnts.gnlnac_ln_intrate; the snapshot holds
        product 16 only while nsecmsmemap maps product 13, so the snapshot yields
        nothing. The undated source must be disclosed as a note."""
        assert report["provenance"]["part8a_msme"]["status"] == "ok"
        assert "gnlnac_ln_intrate" in report["provenance"]["part8a_msme"]["note"]
        msme = report["part8a_msme"][0]
        assert msme["account_count"] > 0
        assert msme["amount_lakhs"] > 0
        assert 0 < msme["min_interest_rate"] <= msme["weighted_avg_interest_rate"]
        assert msme["weighted_avg_interest_rate"] <= msme["max_interest_rate"]

    def test_coverage_is_disclosed(self, report):
        """genln_rpt_day covers product 16 only; the excluded remainder must be stated."""
        coverage = report["coverage"]
        assert coverage["covered_lakhs"] == report["summary"]["total_loan_book"]
        assert coverage["uncovered_accounts"] > 0
        assert 0 < coverage["covered_pct"] < 100

    def test_owned_funds_reconcile_with_part1(self, report):
        totals = {
            row["particulars"]: row["amount_lakhs"]
            for row in report["part1_capital"]
            if row["gl_group"] == "TOTAL"
        }
        assert totals["Owned Funds"] == pytest.approx(
            totals["Share Capital"] + totals["Reserves and Surplus"], abs=0.05
        )

    def test_different_periods_give_different_figures(self, report):
        """The old generator ignored the period entirely for DB-backed sections."""
        april = get_dnbs02_report_data(frequency="monthly", period="2026-04")
        assert april["summary"]["total_loan_book"] != report["summary"]["total_loan_book"]


@requires_db
class TestExcelExport:
    def test_all_sheets_survive(self, workbook):
        assert len(workbook.sheetnames) == 28

    def test_no_prior_filing_data_reaches_the_export(self, workbook):
        offences = []
        for name in workbook.sheetnames:
            for row in workbook[name].iter_rows():
                for cell in row:
                    text = str(cell.value or "")
                    for marker in PRIOR_FILING_MARKERS:
                        if marker.lower() in text.lower():
                            offences.append(f"{name}!{cell.coordinate}: {marker}")
        assert not offences, "prior filing data leaked: " + "; ".join(offences[:10])

    def test_untouched_sheets_carry_no_reported_data(self, workbook):
        """Sheets the generator has no source for must not carry inherited figures.

        Static RBI instruction text is legitimate template content and stays; what must
        never appear is reported data - figures, PANs or contact details.
        """
        for name in ("AuthorisedSignatory", "DNBS02_Annex3", "DNBS02_Annex12", "DNBS02_PART9"):
            offences = []
            for row in workbook[name].iter_rows(min_row=13):
                for cell in row:
                    if cell.value is None:
                        continue
                    if isinstance(cell.value, (int, float)):
                        offences.append(f"{cell.coordinate}={cell.value!r}")
                    elif PAN_RE.search(str(cell.value)) or EMAIL_RE.search(str(cell.value)):
                        offences.append(f"{cell.coordinate}={cell.value!r}")
            assert not offences, f"{name} carries reported data: {offences[:5]}"

    def test_part1_values_land_on_their_rbi_lines(self, workbook, report):
        """Verified against the label in column B, not against a row number."""
        sheet = workbook["DNBS02_PART1"]
        totals = {
            row["particulars"]: row["amount_lakhs"]
            for row in report["part1_capital"]
            if row["gl_group"] == "TOTAL"
        }
        share_row = _find_label_row(sheet, "2 Share Capital")
        reserves_row = _find_label_row(sheet, "3 Reserves and Surplus")
        assert sheet[f"C{share_row}"].value == pytest.approx(totals["Share Capital"])
        assert sheet[f"C{reserves_row}"].value == pytest.approx(totals["Reserves and Surplus"])

    def test_owned_funds_land_on_the_owned_fund_line(self, workbook, report):
        sheet = workbook["DNBS02_PART4"]
        row = _find_label_row(sheet, "Owned Fund (from Part 1)")
        assert sheet[f"C{row}"].value == pytest.approx(report["summary"]["owned_funds"])

    def test_part8a_rates_land_in_min_max_weighted_columns(self, workbook, report):
        """The old writer wrote a single average into G, which is the Min column."""
        sheet = workbook["DNBS02_PART8A"]
        row = _find_label_row(sheet, "A.1 Direct Exposure")
        msme = report["part8a_msme"][0]
        assert sheet[f"G{row}"].value == pytest.approx(msme["min_interest_rate"])
        assert sheet[f"H{row}"].value == pytest.approx(msme["max_interest_rate"])
        assert sheet[f"I{row}"].value == pytest.approx(msme["weighted_avg_interest_rate"])
        assert sheet[f"C{row}"].value == msme["account_count"]
        assert sheet[f"D{row}"].value == pytest.approx(msme["amount_lakhs"])

    def test_part8c_standard_and_npa_lines(self, workbook, report):
        sheet = workbook["DNBS02_PART8C"]
        standard = _find_label_row(sheet, "(i) Standard assets")
        npas = _find_label_row(sheet, "3 Total NPAs")
        assert sheet[f"C{standard}"].value == pytest.approx(
            report["summary"]["total_loan_book"], abs=0.05
        )
        assert sheet[f"C{npas}"].value == 0.0

    def test_annex9_columns_match_their_headers(self, workbook, report):
        """Each value must sit under the header naming it."""
        sheet = workbook["DNBS02_Annex9"]
        block = next(b for b in TABLE_BLOCKS if b.sheet == "DNBS02_Annex9")
        first = report["annex9_top_borrowers"][0]
        for col in block.columns:
            header = _norm(sheet[f"{col.column}{block.header_row}"].value)
            assert header.startswith(_norm(col.header))
            assert sheet[f"{col.column}{block.first_row}"].value == (
                first[col.field] if first[col.field] != "" else None
            )

    def test_annex13_counts_are_not_in_the_date_columns(self, workbook):
        """The old writer put account counts into 'Opening Date' and 'Closing Date'."""
        sheet = workbook["DNBS02_Annex13"]
        assert _norm(sheet["H12"].value).startswith("opening date")
        assert _norm(sheet["I12"].value).startswith("closing date")
        assert sheet["H13"].value is None
        assert sheet["I13"].value is None
        assert _norm(sheet["K12"].value).startswith("number of loan accounts")
        assert isinstance(sheet["K13"].value, int)

    def test_annex11_is_empty_when_there_are_no_npas(self, workbook):
        sheet = workbook["DNBS02_Annex11"]
        assert all(sheet.cell(13, col).value is None for col in range(2, 14))

    def test_filinginfo_states_the_limitations(self, workbook, report):
        sheet = workbook["FilingInfo"]
        text = " ".join(
            str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None
        )
        assert report["snapshot_date"] in text
        if report["degraded_sections"]:
            assert "no source" in text.lower()
