"""Approved, PostgreSQL-backed values for regulatory fields absent from Oracle.

The operational warehouse cannot prove every RBI disclosure.  This module provides a
controlled escape hatch for those fields: a migration-managed silver table whose rows
carry maker/checker evidence and point at blank cells in the official workbook.  Values
may fill blanks only; calculated or otherwise populated cells can never be overwritten.
"""

from __future__ import annotations

import datetime as dt
from decimal import Decimal
from typing import Any, Dict, List, Optional

from openpyxl.utils.cell import coordinate_to_tuple


class ApprovedReportValueError(ValueError):
    pass


def load_approved_values(cur, report_id: str, reporting_date: str) -> List[Dict[str, Any]]:
    cur.execute("SELECT to_regclass('silver.regulatory_report_values')")
    if not cur.fetchone()[0]:
        return []
    cur.execute(
        """
        SELECT sheet_name, target_cell, value_numeric, value_text, value_date,
               value_boolean, source_document, checker, approved_at
        FROM silver.regulatory_report_values
        WHERE report_id = %s
          AND reporting_date = CAST(%s AS DATE)
          AND approved_at IS NOT NULL
          AND checker IS NOT NULL
          AND effective_from <= CAST(%s AS DATE)
          AND (effective_to IS NULL OR effective_to >= CAST(%s AS DATE))
        ORDER BY sheet_name, target_cell
        """,
        (report_id, reporting_date, reporting_date, reporting_date),
    )
    rows: List[Dict[str, Any]] = []
    for row in cur.fetchall():
        numeric, text, date_value, boolean = row[2:6]
        value = next(
            value for value in (numeric, text, date_value, boolean) if value is not None
        )
        if isinstance(value, Decimal):
            value = float(value)
        elif isinstance(value, dt.date):
            value = value.strftime("%d/%m/%Y")
        rows.append(
            {
                "sheet_name": row[0],
                "target_cell": row[1],
                "value": value,
                "source_document": row[6],
                "checker": row[7],
                "approved_at": row[8].isoformat() if row[8] else None,
            }
        )
    return rows


def load_approved_declaration(cur, report_id: str, reporting_date: str) -> Optional[Dict[str, Any]]:
    cur.execute("SELECT to_regclass('silver.regulatory_report_declarations')")
    if not cur.fetchone()[0]:
        return None
    cur.execute(
        """
        SELECT coverage_status, declaration_text, source_document, checker, approved_at
        FROM silver.regulatory_report_declarations
        WHERE report_id = %s
          AND reporting_date = CAST(%s AS DATE)
          AND approved_at IS NOT NULL
          AND checker IS NOT NULL
          AND effective_from <= CAST(%s AS DATE)
          AND (effective_to IS NULL OR effective_to >= CAST(%s AS DATE))
        ORDER BY approved_at DESC
        LIMIT 1
        """,
        (report_id, reporting_date, reporting_date, reporting_date),
    )
    row = cur.fetchone()
    if not row:
        return None
    return {
        "coverage_status": row[0],
        "declaration_text": row[1],
        "source_document": row[2],
        "checker": row[3],
        "approved_at": row[4].isoformat() if row[4] else None,
    }


def apply_approved_values(wb, values: List[Dict[str, Any]]) -> int:
    """Write approved values into blank template cells without changing their styles."""
    for item in values:
        sheet_name = item["sheet_name"]
        target_cell = item["target_cell"].upper()
        if sheet_name not in wb.sheetnames:
            raise ApprovedReportValueError(f"Approved value names unknown sheet {sheet_name!r}")
        sheet = wb[sheet_name]
        row, column = coordinate_to_tuple(target_cell)
        if row > sheet.max_row or column > sheet.max_column:
            raise ApprovedReportValueError(
                f"Approved value target {sheet_name}!{target_cell} is outside the template"
            )
        cell = sheet[target_cell]
        if cell.value not in (None, ""):
            raise ApprovedReportValueError(
                f"Approved value cannot overwrite {sheet_name}!{target_cell}={cell.value!r}"
            )
        cell.value = item["value"]
    return len(values)
