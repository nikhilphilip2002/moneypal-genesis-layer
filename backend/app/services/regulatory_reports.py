"""Registry and PostgreSQL-backed generators for the RBI DNBS report family.

The workbook files define presentation.  This module discovers rows from their labels
and bucket columns from their headers, then writes only values obtained from ``silver``.
Unsupported sections stay blank and are reported through provenance.
"""

from __future__ import annotations

import datetime as dt
import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl

from app.services.db_schema import db_cursor
from app.services import dnbs02_service


ASSET_DIR = Path(__file__).resolve().parents[1] / "assets"


@dataclass(frozen=True)
class ReportDefinition:
    id: str
    return_code: str
    name: str
    frequency: str
    template: str
    workbook_output: str
    description: str


REPORTS: Dict[str, ReportDefinition] = {
    "dnbs02": ReportDefinition(
        "dnbs02", "R039", "DNBS02 — Important Financial Parameters", "quarterly",
        "DNBS02_Blank_Template.xlsx", "dnbs02", "Capital, assets, income, asset quality and annexures.",
    ),
    "dnbs13": ReportDefinition(
        "dnbs13", "R233", "DNBS13 — Overseas Investment Details", "quarterly",
        "DNBS13_Blank_Template.xlsx", "dnbs13", "Overseas JV/WOS investment and supervision details.",
    ),
    "dnbs4a": ReportDefinition(
        "dnbs4a", "R234", "DNBS4A — Short-Term Dynamic Liquidity", "quarterly",
        "DNBS4A_Blank_Template.xlsx", "dnbs4a", "Quarterly short-term inflows, outflows and mismatch.",
    ),
    "dnbs4b_structural": ReportDefinition(
        "dnbs4b_structural", "R228", "DNBS4B — Structural Liquidity", "monthly",
        "DNBS4B_Blank_Template.xlsx", "dnbs4b", "Monthly structural-liquidity maturity ladder.",
    ),
    "dnbs4b_irs": ReportDefinition(
        "dnbs4b_irs", "R228", "DNBS4B — Interest Rate Sensitivity", "monthly",
        "DNBS4B_Blank_Template.xlsx", "dnbs4b", "Monthly interest-rate sensitivity statement.",
    ),
}


class ReportError(ValueError):
    pass


def _norm(value: Any) -> str:
    return " ".join(str(value or "").split()).casefold()


def _date_only(value: Any) -> str:
    if isinstance(value, dt.datetime):
        value = value.date()
    return value.isoformat() if hasattr(value, "isoformat") else str(value)


def _find_row(ws, label: str, within: Optional[Tuple[int, int]] = None) -> int:
    target = _norm(label)
    lo, hi = within or (1, ws.max_row)
    matches = [
        row for row in range(lo, min(hi, ws.max_row) + 1)
        if _norm(ws.cell(row, 2).value).startswith(target)
    ]
    if len(matches) != 1:
        raise ReportError(f"{ws.title}: label {label!r} matched rows {matches}")
    return matches[0]


def _set_label_value(ws, label: str, value: Any) -> bool:
    target = _norm(label)
    for row in range(1, ws.max_row + 1):
        if _norm(ws.cell(row, 2).value).startswith(target):
            ws.cell(row, 3).value = value
            return True
    return False


def _bucket_columns(ws, header_row: int, expected: int) -> List[int]:
    columns: List[int] = []
    total_candidates = [
        col for col in range(3, ws.max_column + 1)
        if _norm(ws.cell(header_row, col).value) == "total"
    ]
    last_col = total_candidates[0] if total_candidates else ws.max_column + 1
    for col in range(3, last_col):
        text = _norm(ws.cell(header_row, col).value)
        if not text or text in {"total", "non-sensitive", "remarks"}:
            continue
        if any(token in text for token in ("day", "month", "year")):
            columns.append(col)
    if len(columns) != expected:
        raise ReportError(
            f"{ws.title}: expected {expected} maturity headers at row {header_row}, found {columns}"
        )
    return columns


def _header_column(ws, header_row: int, label: str) -> int:
    target = _norm(label)
    matches = [
        col for col in range(3, ws.max_column + 1)
        if _norm(ws.cell(header_row, col).value).startswith(target)
    ]
    if len(matches) != 1:
        raise ReportError(f"{ws.title}: header {label!r} matched columns {matches}")
    return matches[0]


def _write_vector(
    ws,
    label: str,
    values: Sequence[float],
    *,
    header_row: int,
    within: Tuple[int, int],
    total: bool = True,
) -> int:
    row = _find_row(ws, label, within)
    columns = _bucket_columns(ws, header_row, len(values))
    for col, value in zip(columns, values):
        ws.cell(row, col).value = round(value, 2)
    if total:
        ws.cell(row, _header_column(ws, header_row, "Total")).value = round(sum(values), 2)
    return row


def _write_cumulative(ws, label: str, values: Sequence[float], *, header_row: int, within: Tuple[int, int]) -> None:
    running = 0.0
    cumulative: List[float] = []
    for value in values:
        running += value
        cumulative.append(round(running, 2))
    _write_vector(ws, label, cumulative, header_row=header_row, within=within)


def _stamp_report_sheet(ws, end_date: str, *, quarter: bool = False) -> None:
    display = dt.date.fromisoformat(end_date).strftime("%d-%b-%Y").upper()
    ws["B5"] = f"Reporting {'Quarter' if quarter else 'Period'} End Date :{display}"


def _fill_common_workbook_metadata(wb, definition: ReportDefinition, start: str, end: str) -> None:
    for name in ("FilingInfo", "Filing Info"):
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        _set_label_value(ws, "Return Name", definition.name.split("—", 1)[-1].strip())
        _set_label_value(ws, "Return Code", definition.return_code)
        _set_label_value(ws, "Reporting frequency", definition.frequency.capitalize())
        _set_label_value(ws, "Return Reporting Frequency", definition.frequency.capitalize())
        _set_label_value(ws, "Reporting start date", start)
        _set_label_value(ws, "Reporting Period Start Date", start)
        _set_label_value(ws, "Reporting end date", end)
        _set_label_value(ws, "Reporting Period End Date", end)


def _parse_request(
    definition: ReportDefinition,
    frequency: str,
    period: str,
    start_date: Optional[str],
    end_date: Optional[str],
) -> Tuple[str, str]:
    if start_date and end_date:
        return start_date, end_date
    requested_frequency = (frequency or definition.frequency).lower()
    if requested_frequency != definition.frequency:
        raise ReportError(
            f"{definition.name} is {definition.frequency}; got {requested_frequency}."
        )
    try:
        return dnbs02_service.parse_period_range(requested_frequency, period)
    except dnbs02_service.PeriodError as exc:
        raise ReportError(str(exc)) from exc


def _available_alm_dates(cur) -> List[str]:
    cur.execute(
        "SELECT DISTINCT nbfc_ason_date::date "
        "FROM silver.nbfc_alm_main_detail_ii "
        "WHERE nbfc_ason_date IS NOT NULL ORDER BY 1"
    )
    return [_date_only(row[0]) for row in cur.fetchall()]


def _monthly_periods(dates: Iterable[str]) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    for value in dates:
        date = dt.date.fromisoformat(value)
        next_day = date + dt.timedelta(days=1)
        if next_day.month == date.month:
            continue
        out.append({"value": date.strftime("%Y-%m"), "label": date.strftime("%B %Y"), "end_date": value})
    return list(reversed(out))


def _quarterly_periods(dates: Iterable[str]) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    for value in dates:
        date = dt.date.fromisoformat(value)
        if (date.month, date.day) not in {(3, 31), (6, 30), (9, 30), (12, 31)}:
            continue
        if date.month == 6:
            fy, quarter = date.year, "Q1"
        elif date.month == 9:
            fy, quarter = date.year, "Q2"
        elif date.month == 12:
            fy, quarter = date.year, "Q3"
        else:
            fy, quarter = date.year - 1, "Q4"
        out.append({"value": f"{fy}-{quarter}", "label": f"{quarter} FY{str(fy + 1)[-2:]}", "end_date": value})
    return list(reversed(out))


def list_reports() -> List[Dict[str, Any]]:
    return [
        {
            "id": item.id,
            "return_code": item.return_code,
            "name": item.name,
            "frequency": item.frequency,
            "description": item.description,
            "workbook_output": item.workbook_output,
        }
        for item in REPORTS.values()
    ]


def get_periods(report_id: str) -> Dict[str, Any]:
    definition = REPORTS.get(report_id)
    if not definition:
        raise ReportError(f"Unknown report {report_id!r}")
    if report_id == "dnbs02":
        return dnbs02_service.get_reportable_periods()
    with db_cursor() as (_conn, cur):
        dates = _available_alm_dates(cur)
    periods = _monthly_periods(dates) if definition.frequency == "monthly" else _quarterly_periods(dates)
    return {
        definition.frequency: periods,
        "source_dates": dates,
        "note": "Periods are offered only when silver.nbfc_alm_main_detail_ii has an exact period-end date.",
    }


def _cashflows(cur, end_date: str) -> Dict[str, List[float]]:
    """Ten contractual maturity buckets in thousands, deduplicated at source grain."""
    cur.execute(
        """
        WITH source AS (
            SELECT DISTINCT *
            FROM silver.nbfc_alm_main_detail_ii
            WHERE nbfc_ason_date::date = CAST(%s AS DATE)
              AND nbfc_pay_recv = 'R'
        )
        SELECT nbfc_princ_int,
               COALESCE(SUM(nbfc_col1),0)/1000.0,
               COALESCE(SUM(nbfc_col2),0)/1000.0,
               COALESCE(SUM(nbfc_col3),0)/1000.0,
               COALESCE(SUM(nbfc_col4),0)/1000.0,
               COALESCE(SUM(nbfc_col5),0)/1000.0,
               COALESCE(SUM(nbfc_col6),0)/1000.0,
               COALESCE(SUM(nbfc_col7),0)/1000.0,
               COALESCE(SUM(nbfc_col8),0)/1000.0,
               COALESCE(SUM(nbfc_col9),0)/1000.0,
               COALESCE(SUM(nbfc_col10),0)/1000.0
        FROM source
        GROUP BY nbfc_princ_int
        """,
        (end_date,),
    )
    result = {"P": [0.0] * 10, "I": [0.0] * 10}
    for row in cur.fetchall():
        result[str(row[0]).strip()] = [float(value or 0) for value in row[1:]]
    result["total"] = [result["P"][i] + result["I"][i] for i in range(10)]
    return result


def _require_alm_date(cur, end_date: str) -> None:
    cur.execute(
        "SELECT EXISTS (SELECT 1 FROM silver.nbfc_alm_main_detail_ii "
        "WHERE nbfc_ason_date::date = CAST(%s AS DATE))",
        (end_date,),
    )
    if not cur.fetchone()[0]:
        cur.execute("SELECT MAX(nbfc_ason_date)::date FROM silver.nbfc_alm_main_detail_ii")
        latest = cur.fetchone()[0]
        raise ReportError(
            f"No ALM snapshot exists for period end {end_date}; latest is {_date_only(latest) if latest else 'none'}."
        )


def _report_envelope(definition: ReportDefinition, start: str, end: str) -> Dict[str, Any]:
    return {
        "report_id": definition.id,
        "return_code": definition.return_code,
        "name": definition.name,
        "frequency": definition.frequency,
        "start_date": start,
        "end_date": end,
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "source": "PostgreSQL silver layer",
        "status": "draft",
        "provenance": {},
        "summary": {},
    }


def _build_dnbs13(definition: ReportDefinition, start: str, end: str) -> Dict[str, Any]:
    data = _report_envelope(definition, start, end)
    data["status"] = "blocked"
    data["provenance"] = {
        "overseas_investments": {
            "status": "no_source",
            "row_count": 0,
            "error": (
                "No approved silver overseas-investment register or not-applicable declaration exists. "
                "The workbook is generated blank rather than inferring non-applicability."
            ),
        }
    }
    data["summary"] = {"populated_rows": 0, "message": "Awaiting approved DNBS13 applicability data."}
    return data


def _build_alm_report(definition: ReportDefinition, start: str, end: str) -> Dict[str, Any]:
    data = _report_envelope(definition, start, end)
    with db_cursor() as (_conn, cur):
        _require_alm_date(cur, end)
        flows = _cashflows(cur, end)
    data["cashflows_thousands"] = flows
    data["source_date"] = end
    data["status"] = "partial"
    data["provenance"] = {
        "performing_loan_cashflows": {
            "status": "ok",
            "row_count": 2,
            "note": (
                "Exact-row duplicates are removed at the full silver row grain before aggregation. "
                "Amounts are contractual principal and interest receivables in ₹ thousands."
            ),
        },
        "non_loan_sections": {
            "status": "no_source",
            "row_count": 0,
            "error": (
                "The populated ALM fact contains loan receivables only. Unsupported liability, "
                "investment and OBS lines remain blank."
            ),
        },
    }
    data["summary"] = {
        "principal_thousands": round(sum(flows["P"]), 2),
        "interest_thousands": round(sum(flows["I"]), 2),
        "total_cashflows_thousands": round(sum(flows["total"]), 2),
    }
    return data


def get_report_data(
    report_id: str,
    frequency: str = "",
    period: str = "",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> Dict[str, Any]:
    definition = REPORTS.get(report_id)
    if not definition:
        raise ReportError(f"Unknown report {report_id!r}")
    if report_id == "dnbs02":
        return dnbs02_service.get_dnbs02_report_data(
            frequency=frequency or "quarterly", period=period,
            start_date=start_date, end_date=end_date,
        )
    start, end = _parse_request(definition, frequency, period, start_date, end_date)
    if report_id == "dnbs13":
        return _build_dnbs13(definition, start, end)
    return _build_alm_report(definition, start, end)


def _write_dnbs13(wb, definition: ReportDefinition, data: Dict[str, Any]) -> None:
    _fill_common_workbook_metadata(wb, definition, data["start_date"], data["end_date"])
    _stamp_report_sheet(wb["DNBS13"], data["end_date"])


def _write_dnbs4a(wb, definition: ReportDefinition, data: Dict[str, Any]) -> None:
    _fill_common_workbook_metadata(wb, definition, data["start_date"], data["end_date"])
    ws = wb["DNBS4AShortTermDynamicLiquidity"]
    _stamp_report_sheet(ws, data["end_date"], quarter=True)
    flows = data["cashflows_thousands"]
    # DNBS4A collapses the fourth and fifth structural buckets into 1-3 months.
    interest = [flows["I"][0], flows["I"][1], flows["I"][2], flows["I"][3] + flows["I"][4], flows["I"][5]]
    _write_vector(ws, "6 Interest inflow on performing Advances", interest, header_row=10, within=(38, 88))
    _write_vector(ws, "TOTAL INFLOWS", interest, header_row=10, within=(38, 88))
    _write_vector(ws, "C Mismatch", interest, header_row=10, within=(80, 88))
    _write_cumulative(ws, "D Cumulative mismatch", interest, header_row=10, within=(80, 88))


def _write_dnbs4b(wb, definition: ReportDefinition, data: Dict[str, Any]) -> None:
    _fill_common_workbook_metadata(wb, definition, data["start_date"], data["end_date"])
    flows = data["cashflows_thousands"]
    total = flows["total"]

    structural = wb["DNBS4BStructuralLiquidity"]
    _stamp_report_sheet(structural, data["end_date"])
    for label in ("Advances (Performing)", "(ii)  Term Loans", "(a)  Through Regular Payment Schedule"):
        _write_vector(structural, label, total, header_row=10, within=(139, 194))
    _write_vector(structural, "B.  TOTAL INFLOWS", total, header_row=10, within=(190, 194))
    _write_vector(structural, "C.  Mismatch", total, header_row=10, within=(194, 198))
    _write_cumulative(structural, "D.  Cumulative Mismatch", total, header_row=10, within=(194, 198))

    irs = wb["DNBS4BIRS"]
    _stamp_report_sheet(irs, data["end_date"])
    # The silver source provides contractual maturity but no reliable fixed/floating or
    # repricing flag for product 16. Populate the provable parent rows only.
    for label in ("5 Advances (Performing)", "(ii) Term loans"):
        _write_vector(irs, label, total, header_row=10, within=(135, 188))
    _write_vector(irs, "B TOTAL INFLOWS", total, header_row=10, within=(185, 192))
    _write_vector(irs, "C Mismatch", total, header_row=10, within=(185, 192))
    _write_cumulative(irs, "D Cumulative mismatch", total, header_row=10, within=(185, 192))


def generate_report_excel(
    report_id: str,
    frequency: str = "",
    period: str = "",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> bytes:
    definition = REPORTS.get(report_id)
    if not definition:
        raise ReportError(f"Unknown report {report_id!r}")
    if report_id == "dnbs02":
        return dnbs02_service.generate_dnbs02_excel(
            frequency=frequency or "quarterly", period=period,
            start_date=start_date, end_date=end_date,
        )
    data = get_report_data(report_id, frequency, period, start_date, end_date)
    path = ASSET_DIR / definition.template
    if not path.exists():
        raise ReportError(f"Template asset {definition.template} is missing")
    wb = openpyxl.load_workbook(path, data_only=False)
    if report_id == "dnbs13":
        _write_dnbs13(wb, definition, data)
    elif report_id == "dnbs4a":
        _write_dnbs4a(wb, definition, data)
    else:
        _write_dnbs4b(wb, definition, data)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
