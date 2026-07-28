"""DNBS-02 lineage workbook: the filing, annotated with where every figure came from.

Produces a separate workbook from the filing itself. The return that goes to RBI stays
byte-clean; this one mirrors it sheet by sheet, with the completed template on the left
and, to its right, one row per RBI line item stating the source table, the source
columns, the filter, the derivation and the query. Below that sits the full text of every
query the sheet depends on.

Both this module and the report writer read dnbs02_spec.FIELD_SPECS, so the document
cannot describe a mapping the filing does not use.
"""

import copy
import io
import logging
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services import dnbs02_spec as spec
from app.services.dnbs02_service import (
    CellMapError,
    _find_label_row,
    get_dnbs02_report_data,
    get_template_path,
    write_report_into,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Presentation.
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E2F3")
GAP_FILL = PatternFill("solid", fgColor="FCE4D6")
BLANK_FILL = PatternFill("solid", fgColor="F2F2F2")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
MONO = Font(name="Consolas", size=9)
SMALL = Font(size=9)
BOLD = Font(bold=True, size=10)
TITLE = Font(bold=True, size=12, color="1F3864")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MAP_HEADERS = [
    ("Target cell", 14),
    ("RBI line / field", 34),
    ("Value in filing", 16),
    ("Status", 20),
    ("Source table(s)", 30),
    ("Source column(s)", 34),
    ("Filter / binds applied", 40),
    ("Derivation", 52),
    ("Query", 22),
]

STATUS_WRITTEN = "written"
STATUS_NO_SOURCE = "NO SOURCE"
GAP_COLUMN_WIDTH = 2.5


def _set(sheet, row: int, col: int, value: Any, *, font=None, fill=None,
         wrap: bool = False, border: bool = True):
    cell = sheet.cell(row=row, column=col)
    cell.value = value
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if border:
        cell.border = BOX
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")
    return cell


def _copy_template_block(src, dst) -> int:
    """Mirror the filled report sheet into the lineage sheet. Returns its last column."""
    max_col = max(src.max_column, 1)
    for row in src.iter_rows():
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            new = dst.cell(row=cell.row, column=cell.column)
            new.value = cell.value
            if not cell.has_style:
                continue
            # Assigning cell._style across workbooks copies *indices* into the source
            # workbook's style tables, which the destination does not have - openpyxl
            # then fails at save time with an IndexError. Copy the concrete style
            # objects instead; those are self-contained.
            new.font = copy.copy(cell.font)
            new.fill = copy.copy(cell.fill)
            new.border = copy.copy(cell.border)
            new.alignment = copy.copy(cell.alignment)
            new.number_format = cell.number_format
    for rng in src.merged_cells.ranges:
        try:
            dst.merge_cells(str(rng))
        except Exception:  # noqa: BLE001 - a malformed range must not lose the sheet
            logger.debug("lineage: could not mirror merged range %s on %s", rng, src.title)
    for letter, dim in src.column_dimensions.items():
        if dim.width:
            dst.column_dimensions[letter].width = dim.width
    for idx, dim in src.row_dimensions.items():
        if dim.height:
            dst.row_dimensions[idx].height = dim.height
    return max_col


# ---------------------------------------------------------------------------
# Resolving what actually happened to each spec.
# ---------------------------------------------------------------------------


def _section_status(data: Dict[str, Any], section: str) -> Tuple[str, str]:
    """(status, reason) for a section key, from the report's own provenance."""
    prov = (data.get("provenance") or {}).get(section)
    if not prov:
        return ("", "")
    return (prov.get("status", ""), prov.get("note", "") or prov.get("reason", "") or "")


def _resolve_coord(report_wb, fs: spec.FieldSpec) -> str:
    """The workbook coordinate a spec lands on, resolved against the real template."""
    if fs.kind == spec.KIND_META:
        return fs.coord
    if fs.kind == spec.KIND_NO_SOURCE:
        return "-"
    if fs.sheet not in report_wb.sheetnames:
        return "(sheet absent)"
    if fs.kind == spec.KIND_TABLE:
        block = next((b for b in spec.TABLE_BLOCKS if b.sheet == fs.sheet), None)
        if block is None:
            return f"{fs.column}?"
        return f"{fs.column}{block.first_row}:{fs.column}{block.first_row + block.max_rows - 1}"
    try:
        row = _find_label_row(report_wb[fs.sheet], fs.rbi_line, fs.within)
    except CellMapError:
        return f"{fs.column}? (label not found)"
    return f"{fs.column}{row}"


def _spec_value_and_status(
    data: Dict[str, Any], fs: spec.FieldSpec, meta: Dict[str, Any]
) -> Tuple[Any, str, Any]:
    """(displayed value, status text, fill) for one spec."""
    if fs.kind == spec.KIND_NO_SOURCE:
        return ("", STATUS_NO_SOURCE, GAP_FILL)

    if fs.kind == spec.KIND_META:
        return (meta.get(fs.coord, "(see sheet)"), STATUS_WRITTEN, OK_FILL)

    if fs.kind == spec.KIND_TABLE:
        rows = data.get(fs.data_key or fs.section) or []
        status, _reason = _section_status(data, fs.section)
        if rows:
            return (f"{len(rows)} row(s)", STATUS_WRITTEN, OK_FILL)
        return ("", f"blank - section {status or 'empty'}", BLANK_FILL)

    # KIND_LINE
    if fs.gate and not data.get(fs.gate):
        status, _ = _section_status(data, fs.section)
        return ("", f"blank - section {status or 'empty'}", BLANK_FILL)
    value = fs.value(data) if fs.value else None
    if value is None:
        return ("", "blank - no GL head fed this line", BLANK_FILL)
    return (value, STATUS_WRITTEN, OK_FILL)


def _spec_reason(data: Dict[str, Any], fs: spec.FieldSpec) -> str:
    """Why a spec is blank, or the section caveat when it is not."""
    if fs.kind == spec.KIND_NO_SOURCE:
        return fs.no_source_reason
    _status, reason = _section_status(data, fs.section)
    return reason


# ---------------------------------------------------------------------------
# Sheet writers.
# ---------------------------------------------------------------------------


def _write_bindings_sheet(wb, data: Dict[str, Any]) -> None:
    """How the UI's period selection became query parameters.

    This is the sheet to read first when asking "why does this quarter show these
    numbers": every dated figure in the return traces back to one of these four binds.
    """
    sheet = wb.create_sheet("_Bindings", 0)
    sheet.sheet_properties.tabColor = "1F3864"
    b = data.get("bindings") or {}

    _set(sheet, 1, 1, "DNBS-02 - how the UI selection became query parameters",
         font=TITLE, border=False)
    row = 3

    _set(sheet, row, 1, "1. What the user selected", font=BOLD, fill=SUBHEAD_FILL)
    for c in range(2, 6):
        _set(sheet, row, c, "", fill=SUBHEAD_FILL)
    row += 1
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 30
    sheet.column_dimensions["D"].width = 34
    sheet.column_dimensions["E"].width = 62

    for label, value, origin in (
        ("Reporting frequency", data.get("frequency", ""), "UI dropdown (monthly / quarterly / yearly)"),
        ("Reporting period", data.get("period", ""), "UI period picker, restricted to periods the warehouse can back"),
    ):
        _set(sheet, row, 1, label, font=SMALL)
        _set(sheet, row, 2, value, font=BOLD)
        _set(sheet, row, 5, origin, font=SMALL, wrap=True)
        row += 1

    row += 1
    _set(sheet, row, 1, "2. What those selections derive", font=BOLD, fill=SUBHEAD_FILL)
    for c in range(2, 6):
        _set(sheet, row, c, "", fill=SUBHEAD_FILL)
    row += 1
    for col, header in enumerate(("Bind", "Value", "Derived by", "Used as", "Meaning"), start=1):
        _set(sheet, row, col, header, font=HEADER_FONT, fill=HEADER_FILL)
    row += 1

    derivations = [
        ("start_date", b.get("start_date", ""), "parse_period_range(frequency, period)",
         "period start",
         "First day of the selected period. Reported on FilingInfo only - every "
         "point-in-time figure is measured at the period end, not across the range."),
        ("end_date", b.get("end_date", ""), "parse_period_range(frequency, period)",
         "period end",
         "Last day of the selected period. For quarterly, RBI quarters are financial "
         "(Q1 = Apr-Jun, Q4 = Jan-Mar of the following calendar year)."),
        ("snapshot_date", b.get("snapshot_date", ""), "resolve_snapshot_date(end_date)",
         "gnlnr_report_date = %s",
         "Must equal the period end exactly. If bronze.genln_rpt_day holds no snapshot "
         "on that date the report is refused rather than silently reported on the "
         "nearest earlier snapshot."),
        ("gl_year", b.get("gl_year", ""), "_gl_year_for(end_date) = int(end_date[:4])",
         "glbbal_year = %s",
         "bronze.glbbal is keyed by branch and year only - no date dimension - so Parts "
         "1, 3, 4, 6 and Annex 10 are annual figures however short the selected period."),
    ]
    for name, value, derived_by, used_as, meaning in derivations:
        _set(sheet, row, 1, name, font=MONO)
        _set(sheet, row, 2, str(value), font=BOLD)
        _set(sheet, row, 3, derived_by, font=MONO, wrap=True)
        _set(sheet, row, 4, used_as, font=MONO, wrap=True)
        _set(sheet, row, 5, meaning, font=SMALL, wrap=True)
        sheet.row_dimensions[row].height = 46
        row += 1

    row += 1
    _set(sheet, row, 1, "3. Consequence for this run", font=BOLD, fill=SUBHEAD_FILL)
    for c in range(2, 6):
        _set(sheet, row, c, "", fill=SUBHEAD_FILL)
    row += 1
    cov = data.get("coverage") or {}
    notes = [
        f"Portfolio measured as at {b.get('snapshot_date', '')}; "
        f"GL trial balance taken from year {b.get('gl_year', '')}.",
    ]
    if cov.get("uncovered_accounts"):
        notes.append(
            f"Coverage: {cov.get('covered_accounts')} accounts "
            f"({cov.get('covered_lakhs')} lakh, {cov.get('covered_pct')}% of the open book) "
            f"are in the return. {cov.get('uncovered_accounts')} open accounts "
            f"({cov.get('uncovered_lakhs')} lakh) have no dated snapshot in "
            f"bronze.genln_rpt_day and are excluded."
        )
    if data.get("degraded_sections"):
        notes.append("Sections with no data: " + ", ".join(data["degraded_sections"]))
    for note in notes:
        _set(sheet, row, 1, note, font=SMALL, wrap=True)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        sheet.row_dimensions[row].height = 32
        row += 1

    sheet.sheet_view.showGridLines = False


def _write_summary_sheet(wb, data: Dict[str, Any]) -> None:
    """Per-section provenance: what ran, what it returned, and why it did not."""
    sheet = wb.create_sheet("_Sections", 1)
    sheet.sheet_properties.tabColor = "1F3864"
    _set(sheet, 1, 1, "DNBS-02 - section provenance", font=TITLE, border=False)

    widths = (30, 14, 10, 34, 40, 46, 60)
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    row = 3
    headers = ("Section", "Status", "Rows", "Source table(s)", "Filter", "Grain", "Reason / caveat")
    for col, header in enumerate(headers, start=1):
        _set(sheet, row, col, header, font=HEADER_FONT, fill=HEADER_FILL)
    row += 1

    prov = data.get("provenance") or {}
    for key in sorted(prov):
        entry = prov[key]
        source = spec.SOURCES.get(key)
        status = entry.get("status", "")
        fill = OK_FILL if status == "ok" else (GAP_FILL if status == "no_source" else BLANK_FILL)
        _set(sheet, row, 1, key, font=MONO)
        _set(sheet, row, 2, status, fill=fill, font=SMALL)
        _set(sheet, row, 3, entry.get("row_count", 0), font=SMALL)
        _set(sheet, row, 4, source.table if source else "-", font=SMALL, wrap=True)
        _set(sheet, row, 5, source.filters if source else "-", font=SMALL, wrap=True)
        _set(sheet, row, 6, source.grain if source else "-", font=SMALL, wrap=True)
        reason = entry.get("note", "") or ""
        if source and source.caveat:
            reason = (reason + "  " if reason else "") + source.caveat
        _set(sheet, row, 7, reason, font=SMALL, wrap=True)
        sheet.row_dimensions[row].height = 42
        row += 1

    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False


def _write_mapping_block(sheet, report_wb, data: Dict[str, Any],
                         specs: List[spec.FieldSpec], meta: Dict[str, Any],
                         first_col: int) -> None:
    """The mapping table to the right of the template, plus the query appendix below it."""
    for offset, (header, width) in enumerate(MAP_HEADERS):
        col = first_col + offset
        sheet.column_dimensions[get_column_letter(col)].width = width

    row = 1
    _set(sheet, row, first_col, "Field mapping - where each value on this sheet comes from",
         font=TITLE, border=False)
    row = 2
    for offset, (header, _width) in enumerate(MAP_HEADERS):
        _set(sheet, row, first_col + offset, header, font=HEADER_FONT, fill=HEADER_FILL, wrap=True)
    row += 1

    sections_used: List[str] = []
    for fs in specs:
        value, status, fill = _spec_value_and_status(data, fs, meta)
        source = spec.SOURCES.get(fs.section)
        if source is not None and fs.section not in sections_used:
            sections_used.append(fs.section)

        reason = _spec_reason(data, fs)
        derivation = fs.derivation
        if fs.kind == spec.KIND_NO_SOURCE:
            derivation = reason
        elif reason:
            derivation = f"{derivation}\nCaveat: {reason}" if derivation else reason

        binds = ""
        if source is not None:
            binds = source.filters.format(**{k: v for k, v in (data.get("bindings") or {}).items()}) \
                if "{" in source.filters else source.filters

        cells = [
            _resolve_coord(report_wb, fs),
            fs.rbi_line,
            value,
            status,
            source.table if source else ("UI selection" if fs.section == "_bindings" else "-"),
            ", ".join(source.columns) if source else "-",
            binds or "-",
            derivation or "-",
            fs.section if source else "-",
        ]
        for offset, cell_value in enumerate(cells):
            font = MONO if offset in (0, 4, 5, 8) else SMALL
            _set(sheet, row, first_col + offset, cell_value, font=font, wrap=offset >= 4,
                 fill=(fill if offset == 3 else None))
        if fs.unit:
            sheet.cell(row=row, column=first_col + 2).number_format = (
                "#,##0.00" if fs.unit != "count" else "#,##0"
            )
        sheet.row_dimensions[row].height = 40
        row += 1

    # -- Query appendix ------------------------------------------------------
    if not sections_used:
        return
    row += 2
    _set(sheet, row, first_col, "Queries behind this sheet", font=TITLE, border=False)
    row += 1
    for key in sections_used:
        source = spec.SOURCES[key]
        _set(sheet, row, first_col, key, font=BOLD, fill=SUBHEAD_FILL)
        for offset in range(1, len(MAP_HEADERS)):
            _set(sheet, row, first_col + offset, "", fill=SUBHEAD_FILL)
        _set(sheet, row, first_col + 1, source.table, font=SMALL, fill=SUBHEAD_FILL)
        row += 1
        if source.binds:
            _set(sheet, row, first_col, "binds", font=SMALL)
            _set(sheet, row, first_col + 1,
                 ", ".join(f"%s -> {b}" for b in source.binds), font=MONO, wrap=True)
            sheet.merge_cells(start_row=row, start_column=first_col + 1,
                              end_row=row, end_column=first_col + len(MAP_HEADERS) - 1)
            row += 1
        sql = source.sql.strip()
        cell = _set(sheet, row, first_col, sql, font=MONO, wrap=True)
        sheet.merge_cells(start_row=row, start_column=first_col,
                          end_row=row, end_column=first_col + len(MAP_HEADERS) - 1)
        # One line of Consolas 9 is about 12 points; cap so a long CTE does not push the
        # rest of the appendix off the screen.
        sheet.row_dimensions[row].height = min(12.5 * (sql.count("\n") + 2), 420)
        row += 2


def generate_dnbs02_lineage_excel(
    frequency: str = "monthly",
    period: str = "",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> bytes:
    """Build the lineage workbook for one period.

    The filing is generated first and mirrored in, so the values documented here are the
    values that were actually written - not a second, independently computed set.
    """
    data = get_dnbs02_report_data(
        frequency=frequency, period=period, start_date=start_date, end_date=end_date
    )

    report_wb = openpyxl.load_workbook(get_template_path())
    write_report_into(report_wb, data)

    from app.services.dnbs02_service import _filing_meta

    meta = _filing_meta(data)

    out = openpyxl.Workbook()
    default_sheet = out.active
    if default_sheet is not None:
        out.remove(default_sheet)
    _write_bindings_sheet(out, data)
    _write_summary_sheet(out, data)

    # Sheets that carry at least one documented field, in template order.
    for name in report_wb.sheetnames:
        specs = [s for s in spec.FIELD_SPECS if s.sheet == name]
        if name != "FilingInfo" and not specs:
            continue
        src = report_wb[name]
        dst = out.create_sheet(name)
        max_col = _copy_template_block(src, dst)
        # Leave room for the template's own banner text, which sits in column B and runs
        # well past the last column that actually holds a value.
        first_col = max(max_col + 2, 8)
        dst.column_dimensions[get_column_letter(max_col + 1)].width = GAP_COLUMN_WIDTH
        _write_mapping_block(dst, report_wb, data, specs, meta, first_col)
        dst.sheet_view.showGridLines = False

    # Specs whose sheet is a pseudo-name (the B5 period stamp applies to every sheet).
    leftover = [s for s in spec.FIELD_SPECS if s.sheet not in report_wb.sheetnames]
    if leftover:
        dst = out.create_sheet("_AllSheets")
        _set(dst, 1, 1, "Fields written to every DNBS02_* sheet", font=TITLE, border=False)
        _write_mapping_block(dst, report_wb, data, leftover, meta, 1)
        dst.sheet_view.showGridLines = False

    buf = io.BytesIO()
    out.save(buf)
    buf.seek(0)
    return buf.getvalue()
