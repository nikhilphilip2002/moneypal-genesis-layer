"""RBI DNBS-02 return: execution.

The *what* lives in dnbs02_spec (queries, cell map, derivations, documented gaps). This
module is the *how*: it runs the section pipeline against PostgreSQL and renders the
result into the blank template workbook.

Sections used to be fourteen closures inside one 700-line function, communicating
through `nonlocal` scalars and closed-over lists. Dependencies between them were
invisible - Part 4 worked only because Part 1 happened to be defined above it - and no
section could be run or tested on its own. They are now Section objects with declared
inputs, executed by a driver that resolves order and records provenance.
"""

import os
import io
import calendar
import datetime
import logging
from typing import Dict, Any, List, Tuple, Optional

import openpyxl

logger = logging.getLogger(__name__)

# Connection handling and section provenance are shared with the curiosity graph so a
# change to either is made once. Aliased to the private names this module already used.
from app.services.db_schema import (
    get_connection,
    db_cursor as _db_cursor,
    run_section as _run_section,
    SectionResult,
)

from app.services.dnbs02_spec import (  # noqa: F401 - re-exported for callers and tests
    ASSET_CLASS_TO_PART8C_LINE,
    ASSET_CODE_LABELS,
    CellMapError,
    FieldSpec,
    GL_BORROWINGS,
    GL_DESC_TO_PART1_LINE,
    GL_DESC_TO_PART3_LINE,
    GL_FIXED_ASSETS,
    GL_INCOME,
    GL_INVESTMENTS,
    GL_PROVISIONS,
    GL_RESERVES,
    GL_SHARE_CAPITAL,
    KIND_META,
    KIND_NO_SOURCE,
    KIND_TABLE,
    LABEL_COLUMN,
    LineItem,
    NPA_ASSET_CODES,
    PART1_TOTAL_LINES,
    PART2_MATURITY_LINES,
    PART3_INVESTMENT_SCOPE,
    STANDARD_ASSET_CODES,
    Section,
    Source,
    SOURCES,
    TABLE_BLOCKS,
    TableBlock,
    TableColumn,
    FIELD_SPECS,
    _norm,
    part8c_buckets,
)
from app.services import dnbs02_spec as spec

TEMPLATE_FILENAME = "DNBS02_Blank_Template.xlsx"


class PeriodError(ValueError):
    """The requested reporting period is malformed, or has no data behind it."""


def parse_period_range(frequency: str, period: str) -> Tuple[str, str]:
    """Parse a frequency and period code into (start_date, end_date) ISO strings.

    Raises PeriodError on anything malformed. This used to swallow parse failures and
    return hardcoded 2026 dates, so a typo in the period silently produced a May-2026
    report labelled with whatever the caller had asked for.
    """
    freq = (frequency or "").lower().strip()
    p_str = (period or "").strip()
    if not p_str:
        raise PeriodError("A reporting period is required.")

    if freq == "monthly":
        # Expected: YYYY-MM
        try:
            year_s, month_s = p_str.split("-")
            year, month = int(year_s), int(month_s)
            last_day = calendar.monthrange(year, month)[1]
        except Exception as exc:
            raise PeriodError(
                f"Monthly period must look like 'YYYY-MM' (got {period!r})."
            ) from exc
        return f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last_day:02d}"

    if freq == "quarterly":
        # Expected: YYYY-Qn, where the year is the financial year start (Apr-Mar).
        try:
            year_part, q_part = p_str.split("-")
            year = int(year_part.replace("FY", ""))
            q_part = q_part.upper()
        except Exception as exc:
            raise PeriodError(
                f"Quarterly period must look like 'YYYY-Q1'..'YYYY-Q4' (got {period!r})."
            ) from exc
        fiscal_quarters = {
            "Q1": (f"{year:04d}-04-01", f"{year:04d}-06-30"),
            "Q2": (f"{year:04d}-07-01", f"{year:04d}-09-30"),
            "Q3": (f"{year:04d}-10-01", f"{year:04d}-12-31"),
            # Q4 of FY <year> runs Jan-Mar of the following calendar year.
            "Q4": (f"{year + 1:04d}-01-01", f"{year + 1:04d}-03-31"),
        }
        if q_part not in fiscal_quarters:
            raise PeriodError(
                f"Unknown quarter {q_part!r}; expected Q1, Q2, Q3 or Q4."
            )
        return fiscal_quarters[q_part]

    if freq == "yearly":
        # Expected: YYYY-YYYY (financial year) or a single YYYY.
        try:
            if "-" in p_str:
                start_s, end_s = p_str.split("-")
                start_year, end_year = int(start_s.replace("FY", "")), int(end_s)
                if end_year != start_year + 1:
                    raise PeriodError(
                        f"Financial year {period!r} must span consecutive years "
                        f"(e.g. '{start_year}-{start_year + 1}')."
                    )
            else:
                start_year = int(p_str.replace("FY", ""))
                end_year = start_year + 1
        except PeriodError:
            raise
        except Exception as exc:
            raise PeriodError(
                f"Yearly period must look like 'YYYY-YYYY' or 'YYYY' (got {period!r})."
            ) from exc
        return f"{start_year:04d}-04-01", f"{end_year:04d}-03-31"

    raise PeriodError(
        f"Unknown frequency {frequency!r}; expected 'monthly', 'quarterly' or 'yearly'."
    )


def get_available_snapshot_dates(cur: Any) -> List[str]:
    """Month-end dates for which silver.loan_daily_snapshot_summary actually holds a portfolio snapshot."""
    cur.execute(spec.SNAPSHOT_DATES_SQL)
    return [r[0].isoformat() for r in cur.fetchall()]


def get_available_gl_years(cur: Any) -> List[int]:
    """Financial years for which silver.gl_daily_balances holds a trial balance.

    glbbal is keyed by year only, so Parts 1/3/4 cannot be produced at sub-annual
    granularity from this warehouse.
    """
    cur.execute(spec.GL_YEARS_SQL)
    return [int(r[0]) for r in cur.fetchall()]


def resolve_snapshot_date(cur: Any, end_date: str) -> str:
    """Return the snapshot date to report on, or explain why the period is unusable.

    DNBS-02 point-in-time fields are measured as at the period end, so the warehouse must
    hold a snapshot on exactly that date. Silently reporting the nearest earlier snapshot
    would mislabel the return.
    """
    available = get_available_snapshot_dates(cur)
    if not available:
        raise PeriodError("silver.loan_daily_snapshot_summary holds no portfolio snapshots at all.")
    if end_date in available:
        return end_date
    raise PeriodError(
        f"No portfolio snapshot exists for period end {end_date}. "
        f"silver.loan_daily_snapshot_summary holds month-end snapshots for: {', '.join(available)}."
    )


def _f(value: Any) -> float:
    """Coerce a DB numeric (Decimal, None) to float."""
    return float(value or 0)


def get_reportable_periods() -> Dict[str, Any]:
    """Periods the warehouse can actually back, so the UI cannot offer unbacked ones.

    The period dropdown used to list fixed options regardless of what the database held;
    picking an unbacked one silently produced fabricated figures.
    """
    with _db_cursor() as (_conn, cur):
        snapshots = get_available_snapshot_dates(cur)
        gl_years = get_available_gl_years(cur)
    monthly = [
        {"value": s[:7], "label": datetime.date.fromisoformat(s).strftime("%B %Y"), "end_date": s}
        for s in snapshots
    ]
    return {
        "monthly": list(reversed(monthly)),
        "snapshot_dates": snapshots,
        "gl_years": gl_years,
        # Quarterly and yearly returns need a period-end snapshot on the quarter/year end.
        "note": (
            "Point-in-time sections require a silver.loan_daily_snapshot_summary snapshot dated exactly "
            "on the period end. Parts 1, 3 and 4 come from silver.gl_daily_balances, which is keyed "
            "by year only and therefore cannot be produced at sub-annual granularity."
        ),
    }


def _gl_year_for(end_date: str) -> int:
    """Calendar year used to select a trial balance from silver.gl_daily_balances.

    glbbal is keyed by branch and year only - there is no date dimension - so Parts 1,
    3 and 4 cannot be produced at sub-annual granularity from this warehouse.
    """
    return int(end_date[:4])


# ---------------------------------------------------------------------------
# Pipeline context and stages.
# ---------------------------------------------------------------------------


class Ctx:
    """Everything a section reads and writes.

    Replaces the `nonlocal` scalars and closed-over lists of the previous builder. A
    section takes a Ctx and returns a row count; nothing else is shared, so a section can
    be executed on its own with a hand-built Ctx.
    """

    def __init__(self, cur, conn, start_date: str, end_date: str, snapshot_date: str,
                 gl_year: int, gl_years: List[int]):
        self.cur = cur
        self.conn = conn
        self.start_date = start_date
        self.end_date = end_date
        self.snapshot_date = snapshot_date
        self.gl_year = gl_year
        self.gl_years = gl_years
        self.gl_available = gl_year in gl_years

        # Section outputs, keyed exactly as they appear in the returned report.
        self.rows: Dict[str, List[Dict[str, Any]]] = {
            "part1_capital": [],
            "part2_loans": [],
            "part2_maturity": [],
            "part3_income": [],
            "part4_nof": [],
            "part6_sensitive": [],
            "part8_asset_quality": [],
            "part8a_msme": [],
            "annex2_shareholders": [],
            "annex9_top_borrowers": [],
            "annex10_top_investments": [],
            "annex11_top_npas": [],
            "annex13_branches": [],
        }
        self.coverage: Dict[str, Any] = {}
        self.totals: Dict[str, Any] = {
            "total_loan_book": 0.0,
            "accrued_interest": 0.0,
            "provision_held": 0.0,
            "npa_amount": 0.0,
            "owned_funds": 0.0,
            "account_count": 0,
            "borrower_count": 0,
        }

    @property
    def bindings(self) -> Dict[str, Any]:
        """The query binds derived from the UI's period selection."""
        return {
            "start_date": self.start_date,
            "end_date": self.end_date,
            "snapshot_date": self.snapshot_date,
            "gl_year": self.gl_year,
        }


def _gl_reason(ctx: Ctx) -> str:
    return (
        f"silver.gl_daily_balances has no trial balance for year {ctx.gl_year} "
        f"(available: {ctx.gl_years})."
    )


def _sec_summary(ctx: Ctx) -> int:
    ctx.cur.execute(spec.SUMMARY_SQL, (ctx.snapshot_date,))
    row = ctx.cur.fetchone()
    if not row or not row[0]:
        return 0
    ctx.totals["account_count"] = int(row[0])
    ctx.totals["borrower_count"] = int(row[1])
    ctx.totals["total_loan_book"] = round(_f(row[2]), 2)
    ctx.totals["accrued_interest"] = round(_f(row[3]), 2)
    ctx.totals["provision_held"] = round(_f(row[4]), 2)
    return ctx.totals["account_count"]


def _sec_coverage(ctx: Ctx) -> int:
    # silver.loan_daily_snapshot_summary is the only table with a genuine as-of dimension, but it does
    # not cover the whole book: it holds product 16 only, so products 1 and 13 have no
    # dated snapshot. Falling back to silver.loan_account_master for those would reintroduce
    # undated balances reported as if they were period-end figures, so the uncovered
    # portion is disclosed instead of silently dropped or back-filled.
    ctx.cur.execute(spec.COVERAGE_SQL)
    row = ctx.cur.fetchone()
    cov = ctx.coverage
    cov["uncovered_accounts"] = int(row[0] or 0)
    cov["uncovered_lakhs"] = round(_f(row[1]), 2)
    cov["covered_accounts"] = ctx.totals["account_count"]
    cov["covered_lakhs"] = ctx.totals["total_loan_book"]
    denom = ctx.totals["total_loan_book"] + cov["uncovered_lakhs"]
    cov["covered_pct"] = (
        round(ctx.totals["total_loan_book"] / denom * 100, 2) if denom else 0.0
    )
    return 1


def _sec_part1(ctx: Ctx) -> int:
    ctx.cur.execute(
        spec.PART1_SQL, (ctx.gl_year, GL_SHARE_CAPITAL, GL_RESERVES, GL_BORROWINGS)
    )
    out = ctx.rows["part1_capital"]
    share_capital = reserves = borrowings = 0.0
    for gl_group, descn, amount in ctx.cur.fetchall():
        amt = round(_f(amount), 2)
        if gl_group == GL_SHARE_CAPITAL:
            share_capital += amt
        elif gl_group == GL_RESERVES:
            reserves += amt
        else:
            borrowings += amt
        out.append(
            {"gl_group": gl_group, "particulars": (descn or "").strip(), "amount_lakhs": amt}
        )
    ctx.totals["owned_funds"] = round(share_capital + reserves, 2)
    if out:
        out.append({"gl_group": "TOTAL", "particulars": "Share Capital", "amount_lakhs": round(share_capital, 2)})
        out.append({"gl_group": "TOTAL", "particulars": "Reserves and Surplus", "amount_lakhs": round(reserves, 2)})
        out.append({"gl_group": "TOTAL", "particulars": "Borrowings", "amount_lakhs": round(borrowings, 2)})
        out.append({"gl_group": "TOTAL", "particulars": "Owned Funds", "amount_lakhs": ctx.totals["owned_funds"]})
    return len(out)


def _sec_part2(ctx: Ctx) -> int:
    ctx.cur.execute(spec.PART2_SQL, (ctx.snapshot_date,))
    rows = ctx.cur.fetchall()
    total = sum(_f(r[2]) for r in rows) or 1.0
    for category, cnt, amount in rows:
        ctx.rows["part2_loans"].append(
            {
                "category": category,
                "account_count": int(cnt),
                "amount_lakhs": round(_f(amount), 2),
                "share_pct": round(_f(amount) / total * 100, 2),
            }
        )
    return len(ctx.rows["part2_loans"])


def _sec_part2_maturity(ctx: Ctx) -> int:
    ctx.cur.execute(
        spec.PART2_MATURITY_SQL, (ctx.snapshot_date, ctx.snapshot_date, ctx.snapshot_date)
    )
    for bucket, cnt, amount in ctx.cur.fetchall():
        ctx.rows["part2_maturity"].append(
            {"bucket": bucket, "account_count": int(cnt), "amount_lakhs": round(_f(amount), 2)}
        )
    return len(ctx.rows["part2_maturity"])


def _sec_part3(ctx: Ctx) -> int:
    ctx.cur.execute(spec.PART3_SQL, (ctx.gl_year, GL_INCOME))
    for descn, amount in ctx.cur.fetchall():
        ctx.rows["part3_income"].append(
            {"head": (descn or "").strip(), "amount_lakhs": round(_f(amount), 2)}
        )
    return len(ctx.rows["part3_income"])


def _sec_part4(ctx: Ctx) -> int:
    ctx.rows["part4_nof"].append(
        {"particulars": "Owned Fund (from Part 1)", "amount_lakhs": ctx.totals["owned_funds"]}
    )
    return len(ctx.rows["part4_nof"])


def _sec_part6(ctx: Ctx) -> int:
    ctx.cur.execute(spec.PART6_SQL, (ctx.gl_year, GL_INVESTMENTS))
    for descn, amount in ctx.cur.fetchall():
        label = (descn or "").strip()
        upper = label.upper()
        if "PROPPERT" in upper or "PROPERT" in upper:
            sector = "Real Estate"
        elif "SHARE" in upper or "MUTUAL" in upper:
            sector = "Capital Market"
        else:
            sector = "Other"
        ctx.rows["part6_sensitive"].append(
            {"sector": sector, "particulars": label, "exposure_lakhs": round(_f(amount), 2)}
        )
    return len(ctx.rows["part6_sensitive"])


def _sec_part8(ctx: Ctx) -> int:
    ctx.cur.execute(spec.PART8_SQL, (ctx.snapshot_date,))
    npa_amount = 0.0
    for asset_code, cnt, amount, provision in ctx.cur.fetchall():
        code = (asset_code or "").strip().upper()
        is_npa = code in NPA_ASSET_CODES
        if is_npa:
            npa_amount += _f(amount)
        ctx.rows["part8_asset_quality"].append(
            {
                "asset_code": code,
                "status": ASSET_CODE_LABELS.get(code, f"Unmapped asset code ({code})"),
                "is_npa": is_npa,
                "count": int(cnt),
                "amount_lakhs": round(_f(amount), 2),
                # Provision comes from the ledger, not from an assumed rate.
                "provision_lakhs": round(_f(provision), 2),
            }
        )
    ctx.totals["npa_amount"] = round(npa_amount, 2)
    return len(ctx.rows["part8_asset_quality"])


def _sec_part8a(ctx: Ctx) -> int:
    ctx.cur.execute(spec.PART8A_SQL)
    row = ctx.cur.fetchone()
    if not row or not row[0]:
        return 0
    ctx.rows["part8a_msme"].append(
        {
            # MSMED Micro/Small/Medium classification needs investment and turnover,
            # which this warehouse does not hold. Only the aggregate MSME exposure and
            # its rate spread are derivable.
            "category": "Micro, Small and Medium Enterprises (aggregate)",
            "account_count": int(row[0]),
            "amount_lakhs": round(_f(row[1]), 2),
            "min_interest_rate": round(_f(row[2]), 2),
            "max_interest_rate": round(_f(row[3]), 2),
            "weighted_avg_interest_rate": round(_f(row[4]), 2),
        }
    )
    return len(ctx.rows["part8a_msme"])


def _sec_annex9(ctx: Ctx) -> int:
    # Aggregated by borrower, not by account: RBI asks for the top 25 *borrowers*, and a
    # borrower may hold several loans.
    ctx.cur.execute(spec.ANNEX9_SQL, (ctx.snapshot_date,))
    for row in ctx.cur.fetchall():
        sanctioned = round(_f(row[4]), 2)
        disbursed = round(_f(row[5]), 2)
        code = (row[9] or "").strip().upper()
        ctx.rows["annex9_top_borrowers"].append(
            {
                "cust_id": str(row[0]),
                "borrower_name": (row[1] or "").strip(),
                # PAN is carried on the snapshot itself; no need to invent "NA".
                "pan": row[2] or "",
                # Legal constitution is not derivable - see the annex9_borrower_type gap
                # in dnbs02_spec.FIELD_SPECS.
                "borrower_type": "",
                "account_count": int(row[3]),
                "sanctioned_amt": sanctioned,
                "disbursed_amt": disbursed,
                "undisbursed_amt": round(max(sanctioned - disbursed, 0.0), 2),
                "principal_outstanding": round(_f(row[6]), 2),
                "accrued_interest": round(_f(row[7]), 2),
                "account_status": ASSET_CODE_LABELS.get(code, code),
                "total_outstanding": round(_f(row[8]), 2),
            }
        )
    return len(ctx.rows["annex9_top_borrowers"])


def _sec_annex10(ctx: Ctx) -> int:
    # Only aggregate GL lines exist. The entity-level detail Annex 10 asks for (name,
    # PAN, nature, group-company flag) has no source, so those fields stay blank rather
    # than being filled with plausible names.
    ctx.cur.execute(spec.ANNEX10_SQL, (ctx.gl_year, GL_INVESTMENTS))
    for descn, amount in ctx.cur.fetchall():
        label = (descn or "").strip()
        upper = label.upper()
        if "MUTUAL" in upper:
            inv_type = "MUTUAL FUNDS"
        elif "SHARE" in upper:
            inv_type = "EQUITY SHARES"
        elif "PROPPERT" in upper or "PROPERT" in upper:
            inv_type = "IMMOVABLE PROPERTY"
        else:
            inv_type = ""
        ctx.rows["annex10_top_investments"].append(
            {
                "entity_name": "",
                "gl_head": label,
                "nature": "",
                "investment_type": inv_type,
                "pan": "",
                "book_value": round(_f(amount), 2),
                "is_group_company": "",
                "amt_outstanding": round(_f(amount), 2),
            }
        )
    return len(ctx.rows["annex10_top_investments"])


def _sec_annex11(ctx: Ctx) -> int:
    ctx.cur.execute(spec.ANNEX11_SQL, (ctx.snapshot_date, list(NPA_ASSET_CODES)))
    for row in ctx.cur.fetchall():
        ctx.rows["annex11_top_npas"].append(
            {
                "borrower_name": row[0] or "",
                "pan": row[1] or "",
                "borrower_type": "",
                "principal_os": round(_f(row[2]), 2),
                "int_due": round(_f(row[3]), 2),
                "asset_code": (row[4] or "").strip(),
                "npa_date": row[5].isoformat() if row[5] else "",
                "last_payment_date": row[6].isoformat() if row[6] else "",
                "sanctioned_amt": round(_f(row[7]), 2),
            }
        )
    return len(ctx.rows["annex11_top_npas"])


def _sec_annex13(ctx: Ctx) -> int:
    ctx.cur.execute(spec.ANNEX13_SQL, (ctx.snapshot_date,))
    for brn_code, customers, accounts, amount in ctx.cur.fetchall():
        code = str(int(brn_code))
        ctx.rows["annex13_branches"].append(
            {
                "branch_code": code,
                # No branch master exists - see the annex13_branch_geography gap in
                # dnbs02_spec.FIELD_SPECS.
                "branch_name": f"Branch {code}",
                "address": "",
                "city": "",
                "state": "",
                "district": "",
                "customer_count": int(customers),
                "account_count": int(accounts),
                "total_outstanding": round(_f(amount), 2),
            }
        )
    return len(ctx.rows["annex13_branches"])


def _gl_available(ctx: Ctx) -> bool:
    return ctx.gl_available


SECTIONS: List[Section] = [
    Section("summary", source=SOURCES["summary"], run=_sec_summary),
    Section("coverage", source=SOURCES["coverage"], requires=("summary",), run=_sec_coverage),
    Section(
        "part1_capital",
        source=SOURCES["part1_capital"],
        run=_sec_part1,
        precondition=_gl_available,
        precondition_reason=_gl_reason,
    ),
    Section("part2_loans", source=SOURCES["part2_loans"], run=_sec_part2),
    Section("part2_maturity", source=SOURCES["part2_maturity"], run=_sec_part2_maturity),
    Section(
        "part3_income",
        source=SOURCES["part3_income"],
        run=_sec_part3,
        precondition=_gl_available,
        precondition_reason=_gl_reason,
    ),
    Section(
        "part3_expenses",
        no_source_reason=(
            "No reliable GL-head to RBI-line mapping for expense accounts; "
            "extgl classification flags are NULL on all rows."
        ),
    ),
    Section(
        "part4_nof",
        source=SOURCES["part4_nof"],
        requires=("part1_capital",),
        run=_sec_part4,
        precondition=lambda ctx: ctx.gl_available and bool(ctx.totals["owned_funds"]),
        precondition_reason=lambda ctx: (
            "Depends on Part 1, which has no trial balance for this period."
        ),
    ),
    Section(
        "part6_sensitive",
        source=SOURCES["part6_sensitive"],
        run=_sec_part6,
        precondition=_gl_available,
        precondition_reason=lambda ctx: (
            f"silver.gl_daily_balances has no trial balance for year {ctx.gl_year} "
            f"(available: {ctx.gl_years})."
        ),
    ),
    Section("part8_asset_quality", source=SOURCES["part8_asset_quality"], run=_sec_part8),
    Section(
        "part8a_msme",
        source=SOURCES["part8a_msme"],
        run=_sec_part8a,
        note=(
            "Sourced from silver.loan_account_master (rate: gnlnac_ln_intrate) because "
            "nsecmsmemap maps product-13 accounts, which have no dated snapshot. "
            "genlnacnts has no as-of dimension, so the outstanding amount is a "
            "current balance rather than a period-end one."
        ),
    ),
    Section(
        "part8a_msme_size_split",
        no_source_reason=(
            "MSMED Micro/Small/Medium classification requires investment in plant and "
            "machinery and turnover; silver.msme_sector_classification_mapping carries only collateral value and LTV."
        ),
    ),
    Section(
        "annex2_shareholders",
        no_source_reason=(
            "No share register in the warehouse (silver.customer_share_holdings does not exist)."
        ),
    ),
    Section("annex9_top_borrowers", source=SOURCES["annex9_top_borrowers"], run=_sec_annex9),
    Section(
        "annex10_investment_totals",
        source=SOURCES["annex10_investment_totals"],
        run=_sec_annex10,
        precondition=_gl_available,
        precondition_reason=_gl_reason,
    ),
    Section(
        "annex10_investment_entities",
        no_source_reason=(
            "No entity-level investment register in the warehouse; silver.gl_daily_balances carries "
            "only aggregate investment GL heads, with no counterparty name or PAN."
        ),
    ),
    Section("annex11_top_npas", source=SOURCES["annex11_top_npas"], run=_sec_annex11),
    Section("annex13_branches", source=SOURCES["annex13_branches"], run=_sec_annex13),
    Section(
        "annex13_branch_geography",
        no_source_reason=(
            "No branch master; customer_kyc_details district/state columns are 100% NULL "
            "and gnlnr_adh_district holds numeric codes with no reference table."
        ),
    ),
]

SECTIONS_BY_KEY: Dict[str, Section] = {s.key: s for s in SECTIONS}


def _run_pipeline(ctx: Ctx, provenance: Dict[str, Dict[str, Any]]) -> None:
    """Execute every section in declared order, recording why each produced what it did.

    A section is skipped - and says so - when its precondition fails or when a section it
    requires did not produce rows. Nothing here swallows a failure: run_section records
    an error and rolls the connection back so one bad section cannot kill the return.
    """
    for section in SECTIONS:
        if section.run is None:
            provenance[section.key] = SectionResult(
                section.key, "no_source", 0, section.no_source_reason
            ).as_dict()
            continue

        unmet = [
            dep
            for dep in section.requires
            if provenance.get(dep, {}).get("status") not in ("ok",)
        ]
        if unmet:
            provenance[section.key] = SectionResult(
                section.key,
                "no_source",
                0,
                f"Depends on section(s) {', '.join(unmet)}, which produced no rows.",
            ).as_dict()
            continue

        if section.precondition is not None and not section.precondition(ctx):
            reason = section.precondition_reason
            provenance[section.key] = SectionResult(
                section.key,
                "no_source",
                0,
                reason(ctx) if callable(reason) else (reason or ""),
            ).as_dict()
            continue

        run = section.run
        _run_section(
            provenance,
            section.key,
            lambda fn=run: fn(ctx),
            ctx.conn,
            note=section.note,
        )


def get_dnbs02_report_data(
    frequency: str = "monthly",
    period: str = "",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> Dict[str, Any]:
    """Build the RBI DNBS-02 return from PostgreSQL.

    Every figure is either sourced from the warehouse or left empty. There is no
    fallback data: a section with no feed reports status "no_source" and renders blank,
    because a plausible-looking invented number in a regulatory return is worse than a
    gap.
    """
    if not start_date or not end_date:
        calc_start, calc_end = parse_period_range(frequency, period)
        start_date = start_date or calc_start
        end_date = end_date or calc_end

    try:
        d1 = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        d2 = datetime.datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError as exc:
        raise PeriodError(f"Dates must be ISO YYYY-MM-DD (got {start_date!r}..{end_date!r}).") from exc
    if d2 < d1:
        raise PeriodError(f"Period end {end_date} precedes period start {start_date}.")
    num_days = (d2 - d1).days + 1

    provenance: Dict[str, Dict[str, Any]] = {}

    with _db_cursor() as (conn, cur):
        snapshot_date = resolve_snapshot_date(cur, end_date)
        ctx = Ctx(
            cur=cur,
            conn=conn,
            start_date=start_date,
            end_date=end_date,
            snapshot_date=snapshot_date,
            gl_year=_gl_year_for(end_date),
            gl_years=get_available_gl_years(cur),
        )
        _run_pipeline(ctx, provenance)

    if ctx.coverage.get("uncovered_accounts"):
        logger.warning(
            "DNBS-02 %s: %s open accounts (%.2f lakh) have no dated snapshot in "
            "silver.loan_daily_snapshot_summary and are excluded from the return.",
            ctx.snapshot_date,
            ctx.coverage["uncovered_accounts"],
            ctx.coverage["uncovered_lakhs"],
        )

    total_loan_book = ctx.totals["total_loan_book"]
    npa_amount = ctx.totals["npa_amount"]
    gross_npa_pct = round(npa_amount / total_loan_book * 100, 2) if total_loan_book else 0.0

    # A section is "live" only if its own query actually ran and returned rows - not
    # merely because the connection opened.
    live_sections = sorted(k for k, v in provenance.items() if v["status"] == "ok")
    degraded_sections = sorted(k for k, v in provenance.items() if v["status"] != "ok")

    return {
        "frequency": frequency,
        "period": period,
        "start_date": start_date,
        "end_date": end_date,
        "snapshot_date": ctx.snapshot_date,
        "gl_year": ctx.gl_year,
        "duration_days": num_days,
        "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "provenance": provenance,
        "coverage": ctx.coverage,
        "live_sections": live_sections,
        "degraded_sections": degraded_sections,
        "is_live_pg": bool(live_sections) and not degraded_sections,
        "summary": {
            "total_loan_book": total_loan_book,
            "accrued_interest": ctx.totals["accrued_interest"],
            "account_count": ctx.totals["account_count"],
            "borrower_count": ctx.totals["borrower_count"],
            "owned_funds": ctx.totals["owned_funds"],
            "provision_held": ctx.totals["provision_held"],
            "gross_npa_amount": npa_amount,
            "gross_npa_pct": gross_npa_pct,
            # CRAR needs risk-weighted assets (Part 9), which has no source here. It was
            # previously reported as `24.8 + date_scale_factor * 0.1`.
            "crar_pct": None,
        },
        "part1_capital": ctx.rows["part1_capital"],
        "part2_loans": ctx.rows["part2_loans"],
        "part2_maturity": ctx.rows["part2_maturity"],
        "part3_income": ctx.rows["part3_income"],
        "part4_nof": ctx.rows["part4_nof"],
        "part6_sensitive": ctx.rows["part6_sensitive"],
        "part8_asset_quality": ctx.rows["part8_asset_quality"],
        "part8a_msme": ctx.rows["part8a_msme"],
        "annex2_shareholders": ctx.rows["annex2_shareholders"],
        "annex9_top_borrowers": ctx.rows["annex9_top_borrowers"],
        "annex10_top_investments": ctx.rows["annex10_top_investments"],
        "annex11_top_npas": ctx.rows["annex11_top_npas"],
        "annex13_branches": ctx.rows["annex13_branches"],
        # Bindings are echoed so the lineage workbook (and any caller auditing a figure)
        # can see how the UI's period selection became query parameters.
        "bindings": ctx.bindings,
    }


# ---------------------------------------------------------------------------
# Workbook rendering.
# ---------------------------------------------------------------------------


def get_template_path() -> str:
    """Locate the blank DNBS-02 workbook.

    Only the blank is acceptable. The previous implementation fell back to
    DNBS02_Template.xlsx and then to a source document under docs/, both of which are
    completed filings; any sheet the generator did not overwrite was exported carrying
    the prior filer's data. Regenerate the blank with
    backend/scripts/build_dnbs02_blank_template.py.
    """
    candidates = [
        os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "assets", TEMPLATE_FILENAME)),
        "/srv/backend/app/assets/" + TEMPLATE_FILENAME,
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    raise FileNotFoundError(
        f"Blank RBI DNBS-02 template {TEMPLATE_FILENAME!r} not found (searched {candidates}). "
        "Generate it with backend/scripts/build_dnbs02_blank_template.py."
    )


def _find_label_row(sheet, label: str, within: Optional[Tuple[int, int]] = None) -> int:
    """Row whose label column starts with `label`. Must match exactly one row."""
    target = _norm(label)
    lo, hi = within if within else (1, sheet.max_row)
    matches = [
        r
        for r in range(lo, min(hi, sheet.max_row) + 1)
        if _norm(sheet.cell(r, LABEL_COLUMN).value).startswith(target)
    ]
    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise CellMapError(f"{sheet.title}: no row whose label starts with {label!r}")
    raise CellMapError(
        f"{sheet.title}: label {label!r} is ambiguous, matches rows {matches}. "
        "Use a longer fragment or scope it with `within`."
    )


def validate_cell_map(wb) -> None:
    """Check every declared table header exists before writing anything.

    A template whose columns have shifted must fail here rather than silently writing
    outstanding balances into a date column, which is what the old positional writer did
    on Annex 11 and Annex 13.
    """
    problems: List[str] = []

    for block in TABLE_BLOCKS:
        if block.sheet not in wb.sheetnames:
            problems.append(f"{block.sheet}: sheet missing from template")
            continue
        sheet = wb[block.sheet]
        for col in block.columns:
            actual = _norm(sheet[f"{col.column}{block.header_row}"].value)
            if not actual.startswith(_norm(col.header)):
                problems.append(
                    f"{block.sheet}!{col.column}{block.header_row}: expected header "
                    f"{col.header!r} for field {col.field!r}, found {actual!r}"
                )

    if problems:
        raise CellMapError(
            "DNBS-02 template does not match the declared cell map:\n  "
            + "\n  ".join(problems)
        )


DEFAULT_COLUMN_WIDTH = 8.43
DEFAULT_ROW_HEIGHT = 15.0
MAX_ROW_HEIGHT = 220.0


def _effective_width(sheet, cell) -> float:
    """Usable width for a cell, summing the columns it spans if it is merged."""
    from openpyxl.utils import get_column_letter

    first_col, last_col = cell.column, cell.column
    for rng in sheet.merged_cells.ranges:
        if rng.min_row <= cell.row <= rng.max_row and rng.min_col <= cell.column <= rng.max_col:
            first_col, last_col = rng.min_col, rng.max_col
            break
    total = 0.0
    for col in range(first_col, last_col + 1):
        dim = sheet.column_dimensions.get(get_column_letter(col))
        total += (dim.width if dim and dim.width else DEFAULT_COLUMN_WIDTH)
    return max(total, 1.0)


def _fit_row_height(sheet, cell, text: str) -> None:
    """Grow the row so wrapped text is fully visible.

    Turning on wrap_text without adjusting the row height is what made long values
    overlap the rows beneath them: the text rewraps onto several lines but the row stays
    one line tall, so it renders clipped and bleeds over the labels below.
    """
    width = _effective_width(sheet, cell)
    lines = 0
    for paragraph in str(text).split("\n"):
        # ~1.1 characters per width unit at the default font.
        chars_per_line = max(int(width * 1.1), 1)
        lines += max(1, -(-len(paragraph) // chars_per_line))
    needed = min(DEFAULT_ROW_HEIGHT * lines, MAX_ROW_HEIGHT)
    current = sheet.row_dimensions[cell.row].height or DEFAULT_ROW_HEIGHT
    if needed > current:
        sheet.row_dimensions[cell.row].height = needed


def _safe_set_cell_value(sheet, coord: str, value: Any, wrap_text: bool = True) -> None:
    """Write a cell, skipping merged anchors. Raises on anything unexpected."""
    cell = sheet[coord]
    if type(cell).__name__ == "MergedCell":
        logger.warning("DNBS-02: skipped merged cell %s!%s", sheet.title, coord)
        return
    cell.value = value
    if not isinstance(value, str):
        return

    # The cell may already wrap because the template styles it that way - short values in
    # narrow columns (e.g. "SMA-0 (1-30 days)" in Annex 9 column K) rewrap onto two lines
    # without this module touching the alignment. Any wrapped cell needs its row sized,
    # whoever turned the wrapping on.
    wrapped = bool(cell.alignment and cell.alignment.wrap_text)
    if wrap_text and not wrapped and len(value) > 25:
        from openpyxl.styles import Alignment

        cell.alignment = Alignment(wrap_text=True, vertical="top")
        wrapped = True
    if wrapped:
        _fit_row_height(sheet, cell, value)


def _write_line(sheet, item: LineItem, value: Any) -> None:
    row = _find_label_row(sheet, item.label, item.within)
    _safe_set_cell_value(sheet, f"{item.column}{row}", value, wrap_text=False)


def _write_table(wb, data: Dict[str, Any], block: TableBlock) -> int:
    """Write one annexure table. Rows beyond the data are left blank."""
    if block.sheet not in wb.sheetnames:
        return 0
    sheet = wb[block.sheet]
    rows = data.get(block.source_key) or []
    for idx, record in enumerate(rows[: block.max_rows]):
        row_num = block.first_row + idx
        if block.serial_column:
            _safe_set_cell_value(sheet, f"{block.serial_column}{row_num}", idx + 1, wrap_text=False)
        for col in block.columns:
            value = record.get(col.field, "")
            _safe_set_cell_value(sheet, f"{col.column}{row_num}", value)
    return len(rows[: block.max_rows])


def _log_unmapped_gl_heads(data: Dict[str, Any]) -> None:
    """Report GL heads that reached no RBI line, so a new head is noticed rather than dropped."""
    for row in data.get("part1_capital") or []:
        if row["gl_group"] != "TOTAL" and row["particulars"].upper() not in GL_DESC_TO_PART1_LINE:
            logger.info("DNBS-02 Part 1: GL head %r has no RBI line mapping", row["particulars"])
    for row in data.get("part3_income") or []:
        if row["head"].upper() not in GL_DESC_TO_PART3_LINE:
            logger.info("DNBS-02 Part 3: GL head %r has no RBI line mapping", row["head"])


def _filing_meta(data: Dict[str, Any]) -> Dict[str, Any]:
    """Values for the FilingInfo cells whose coordinates FIELD_SPECS declares."""
    try:
        end_display = datetime.datetime.strptime(data["end_date"], "%Y-%m-%d").strftime("%d/%m/%Y")
        start_display = datetime.datetime.strptime(data["start_date"], "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        start_display, end_display = data["start_date"], data["end_date"]
    return {
        "C11": data["frequency"].capitalize(),
        "C12": start_display,
        "C13": end_display,
        "C15": "LAKHS",
    }


def _filing_remarks(data: Dict[str, Any]) -> List[str]:
    remarks = [
        f"Portfolio snapshot date: {data['snapshot_date']}; "
        f"GL trial balance year: {data['gl_year']}."
    ]
    coverage = data.get("coverage") or {}
    if coverage.get("uncovered_accounts"):
        remarks.append(
            f"Coverage: {coverage['covered_accounts']} accounts "
            f"({coverage['covered_lakhs']} lakh, {coverage['covered_pct']}% of the open "
            f"book). {coverage['uncovered_accounts']} open accounts "
            f"({coverage['uncovered_lakhs']} lakh) have no dated snapshot and are excluded."
        )
    if data.get("degraded_sections"):
        remarks.append(
            "Sections left blank (no source): " + ", ".join(data["degraded_sections"])
        )
    return remarks


def write_report_into(wb, data: Dict[str, Any]) -> None:
    """Write every FIELD_SPEC into an open workbook.

    Driven entirely by dnbs02_spec.FIELD_SPECS: to move a figure to a different RBI line,
    change the spec, and the lineage workbook follows automatically.
    """
    validate_cell_map(wb)
    _log_unmapped_gl_heads(data)

    try:
        end_upper = datetime.datetime.strptime(data["end_date"], "%Y-%m-%d").strftime("%d-%b-%Y").upper()
    except ValueError:
        end_upper = data["end_date"]

    # Period-end stamp on every reporting sheet (FIELD_SPECS declares this as B5).
    for name in wb.sheetnames:
        if name.startswith("DNBS02_"):
            _safe_set_cell_value(wb[name], "B5", f"Reporting Period End Date :{end_upper}", wrap_text=False)

    # -- FilingInfo, including an explicit statement of what has no source ----
    if "FilingInfo" in wb.sheetnames:
        sheet = wb["FilingInfo"]
        _safe_set_cell_value(sheet, "B2", f"Period: {data['start_date']} to {data['end_date']} ({data['frequency']})")
        _safe_set_cell_value(sheet, "B3", f"Generated: {data['generated_at']}")
        for coord, value in _filing_meta(data).items():
            _safe_set_cell_value(sheet, coord, value)
        # The disclosures go in the value column of the template's own "General remarks"
        # row. They previously went into B18/B19/B20, which are inside the sheet's label
        # column and sit directly above the "General remarks" and "Scoping Question"
        # labels, so the text overlapped them.
        try:
            remarks_row = _find_label_row(sheet, "General remarks")
            _safe_set_cell_value(sheet, f"C{remarks_row}", "\n".join(_filing_remarks(data)))
        except CellMapError:
            logger.warning("FilingInfo has no 'General remarks' row; disclosures omitted")

    # -- Every RBI line item, from the registry ------------------------------
    for fs in FIELD_SPECS:
        if fs.kind != spec.KIND_LINE:
            continue
        if fs.sheet not in wb.sheetnames:
            continue
        if fs.gate and not data.get(fs.gate):
            continue
        value = fs.value(data) if fs.value else None
        if value is None:
            # A line the warehouse never fed stays blank, which is not the same as zero.
            continue
        _write_line(wb[fs.sheet], fs.line_item(), value)

    # -- Annexure tables -----------------------------------------------------
    for block in TABLE_BLOCKS:
        written = _write_table(wb, data, block)
        logger.debug("DNBS-02 %s: wrote %d rows", block.sheet, written)


def generate_dnbs02_excel(
    frequency: str = "monthly",
    period: str = "",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> bytes:
    """Render the RBI DNBS-02 return into the blank template workbook.

    Only values with a source are written. Sections the warehouse cannot back are left
    blank and listed on the FilingInfo sheet, rather than being filled with fabricated
    figures.
    """
    data = get_dnbs02_report_data(
        frequency=frequency, period=period, start_date=start_date, end_date=end_date
    )
    wb = openpyxl.load_workbook(get_template_path())
    write_report_into(wb, data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
