"""Derive a blank DNBS-02 workbook from a filed return.

The shipped `DNBS02_Template.xlsx` is not a template: it is a completed GICC filing that
still contains the signatory's email, real borrower and shareholder names, and their PANs.
Because the generator only writes 14 of the 28 sheets, exporting from it leaks the prior
filing's data into every new return.

This script strips every reported value while preserving the sheet structure, the RBI row
taxonomy in column B of the Part sheets, and all formatting, producing the blank the
generator should build on.

Usage:
    uv run python backend/scripts/build_dnbs02_blank_template.py [source.xlsx] [dest.xlsx]
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SRC = REPO_ROOT / "backend" / "app" / "assets" / "DNBS02_Template.xlsx"
DEFAULT_DST = REPO_ROOT / "backend" / "app" / "assets" / "DNBS02_Blank_Template.xlsx"

# Data rows begin immediately below the column-header row, which sits at row 12 on every
# tabular sheet in this workbook.
FIRST_DATA_ROW = 13

# On the Part sheets column B holds RBI's fixed line-item taxonomy ("1 Total Authorized
# Capital", "II Food Credit", ...). That is structure, not reported data, so it stays.
# On the Annex sheets column B is the first data column (serial number or entity name).
PART_SHEET_PREFIX = "DNBS02_PART"

# Free-form sheets whose values live in a labelled column C rather than a table.
KEY_VALUE_SHEETS = {"FilingInfo", "AuthorisedSignatory"}

PAN_RE = re.compile(r"\b[A-Z]{5}\d{4}[A-Z]\b")
EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.]+")


def _clear(sheet, min_row: int, min_col: int, keep_text_subheaders: bool = False) -> int:
    """Blank every non-merged cell at or beyond (min_row, min_col). Returns cells cleared.

    `keep_text_subheaders` preserves text in the first data row. Several Part sheets carry
    a second header tier there - PART8A's Min / Max / Weighted Average under the shared
    "Interest Rate (%)" heading, PART8C's Amount under Balance and Provision Held. Those
    are structure, and clearing them leaves the sheet's columns unlabelled. Part sheets
    report only numbers, so text in that row is always a header.
    """
    cleared = 0
    for row in sheet.iter_rows(
        min_row=min_row, max_row=sheet.max_row, min_col=min_col, max_col=sheet.max_column
    ):
        for cell in row:
            if type(cell).__name__ == "MergedCell":
                continue
            if cell.value is None:
                continue
            if (
                keep_text_subheaders
                and cell.row == FIRST_DATA_ROW
                and isinstance(cell.value, str)
                and cell.value.strip()
            ):
                continue
            cell.value = None
            cleared += 1
    return cleared


def build_blank(src: Path, dst: Path) -> None:
    wb = openpyxl.load_workbook(src)
    total = 0

    for name in wb.sheetnames:
        sheet = wb[name]

        if name in KEY_VALUE_SHEETS:
            # Keep the labels in column B, drop the filed values in column C onward.
            cleared = _clear(sheet, min_row=2, min_col=3)
        elif name.startswith(PART_SHEET_PREFIX):
            # Keep the RBI line-item taxonomy in column B, and any second-tier column
            # headers sitting in the first data row.
            cleared = _clear(sheet, min_row=FIRST_DATA_ROW, min_col=3, keep_text_subheaders=True)
        else:
            # Annex sheets: every column from B is reported data.
            cleared = _clear(sheet, min_row=FIRST_DATA_ROW, min_col=2)

        total += cleared
        print(f"  {name:24s} cleared {cleared:5d} cells")

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    print(f"\nCleared {total} cells -> {dst}")
    verify(dst)


def verify(path: Path) -> None:
    """Fail loudly if any PAN, email or residual value survived the strip."""
    wb = openpyxl.load_workbook(path)
    problems: list[str] = []

    for name in wb.sheetnames:
        sheet = wb[name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value)
                if PAN_RE.search(text):
                    problems.append(f"{name}!{cell.coordinate}: PAN-like value {text!r}")
                if EMAIL_RE.search(text):
                    problems.append(f"{name}!{cell.coordinate}: email {text!r}")
                # Any bare number below the header row is a leftover reported figure.
                if isinstance(cell.value, (int, float)) and cell.row >= FIRST_DATA_ROW:
                    problems.append(f"{name}!{cell.coordinate}: residual number {cell.value!r}")

    if problems:
        print(f"\nVERIFY FAILED - {len(problems)} residual value(s):")
        for p in problems[:40]:
            print("   ", p)
        sys.exit(1)
    print("VERIFY OK - no PANs, emails or residual figures remain.")


if __name__ == "__main__":
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
    dest = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_DST
    print(f"Building blank template from {source}")
    build_blank(source, dest)
