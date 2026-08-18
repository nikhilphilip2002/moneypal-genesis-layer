"""Stage 2 — Extractor: downloaded PDF / CSV / XLSX / TXT -> page-tagged text chunks.

Extraction is **per page** for PDFs. The macro briefs cite sources as
``(document, p.X)`` (see ``genesis_core.rag.DEFAULT_SYSTEM``) and
``app.services.macro`` lifts ``sources[0]["page"]`` into the response envelope, so
flattening a PDF into one blob before chunking silently drops every page citation.

Chunking delegates to ``genesis_core.rag.chunk_text`` so this pipeline's points
land in the same collection with the same chunk geometry as every other ingest.
Spreadsheets and CSVs have no page concept and carry ``page=None``.

Output rows:
    {"path", "document", "source_url", "page", "chunk_index", "text"}
"""
from __future__ import annotations

import csv
import io
import logging
from pathlib import Path

from genesis_core import rag

log = logging.getLogger("macro.extractor")


def _pdf_pages(path: Path) -> list[tuple[int | None, str]]:
    """Return [(page_number, text), ...], skipping pages with no extractable text."""
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    pages: list[tuple[int | None, str]] = []
    for number, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception:
            # A single malformed page should not cost us the rest of the document.
            text = ""
        if text.strip():
            pages.append((number, text))
    return pages


def _spreadsheet_text(path: Path) -> str:
    # openpyxl reads the OOXML formats only. A legacy BIFF .xls raises
    # InvalidFileException, so say what is actually wrong rather than letting an
    # unhelpful parser error surface.
    if path.suffix.lower() == ".xls":
        raise ValueError(
            "legacy .xls (BIFF) is not readable by openpyxl; re-save as .xlsx or add xlrd"
        )

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    blocks: list[str] = []
    for sheet in workbook.worksheets:
        lines: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            if any(values):
                lines.append("\t".join(values))
        if lines:
            blocks.append(f"SHEET: {sheet.title}\n" + "\n".join(lines))
    workbook.close()
    return "\n\n".join(blocks)


def _csv_text(path: Path) -> str:
    try:
        raw = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        raw = path.read_text(encoding="latin-1")
    return "\n".join("\t".join(row) for row in csv.reader(io.StringIO(raw)))


def extract_pages(path: Path) -> list[tuple[int | None, str]]:
    """Return [(page_or_None, text), ...] for one file, dispatched on extension."""
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _pdf_pages(path)
    if suffix in (".xlsx", ".xls"):
        return [(None, _spreadsheet_text(path))]
    if suffix == ".csv":
        return [(None, _csv_text(path))]
    try:
        return [(None, path.read_text(encoding="utf-8", errors="ignore"))]
    except Exception:
        return []


def extract(path: Path) -> list[dict]:
    """Chunk one file into page-tagged rows ready for structuring and embedding.

    One unreadable file must never abort the weekly refresh: a crawl returns whatever a
    government portal chose to publish, including formats no parser here supports. A
    failure is logged and the file skipped, so the remaining sources still ingest and the
    stale-point purge still runs.
    """
    rows: list[dict] = []
    try:
        pages = extract_pages(path)
    except Exception as exc:  # noqa: BLE001 - a bad file is data, not a program error
        log.warning("[extract] unreadable, skipped: %s (%s: %s)", path.name, type(exc).__name__, exc)
        return []
    for page, text in pages:
        if not text.strip():
            continue
        for chunk in rag.chunk_text(text):
            rows.append(
                {
                    "path": str(path),
                    "document": path.name,
                    "source_url": "",
                    "page": page,
                    # Assigned once the whole document is chunked, so the index is
                    # document-scoped rather than page-scoped — point IDs depend on it.
                    "chunk_index": len(rows),
                    "text": chunk,
                }
            )
    if not rows:
        log.info("[extract] no extractable text: %s (scanned PDF?)", path.name)
    else:
        log.info("[extract] %s -> %d chunks", path.name, len(rows))
    return rows


def extract_many(paths: list[Path]) -> list[dict]:
    rows: list[dict] = []
    for path in paths:
        rows.extend(extract(path))
    return rows
