"""Build clean RBI report templates from the filed workbooks supplied by the client.

The source files are examples of completed returns, not safe templates.  This script
preserves their workbook structure and formatting while removing institution, signatory
and reported values.  Runtime report writers fill the clean copies from PostgreSQL.

Usage:
    .venv/bin/python backend/scripts/build_regulatory_templates.py <extracted-directory>
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl


REPO_ROOT = Path(__file__).resolve().parents[2]
ASSET_DIR = REPO_ROOT / "backend" / "app" / "assets"
PAN_RE = re.compile(r"\b[A-Z]{5}\d{4}[A-Z]\b")
EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.]+")

FILES = {
    "DNBS13-Overseas Investment Details (1).xlsx": "DNBS13_Blank_Template.xlsx",
    "DNBS4A-Short Term Dynamic Liquidity (STDL) – Quarterly (3).xlsx": "DNBS4A_Blank_Template.xlsx",
    "DNBS4B-Structural Liquidity & Interest Rate Sensitivity – Monthly.xlsx":
        "DNBS4B_Blank_Template.xlsx",
}


def _blank_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if type(cell).__name__ != "MergedCell":
                cell.value = None


def _blank_key_values(ws, rows: range) -> None:
    for row in rows:
        for col in range(3, ws.max_column + 1):
            cell = ws.cell(row, col)
            if type(cell).__name__ != "MergedCell":
                cell.value = None


def _neutralise_report_header(ws, period_label: str) -> None:
    ws["B4"] = "NBFC Name :"
    ws["B5"] = period_label
    ws["B6"] = "Audit Status :"


def sanitize_dnbs13(wb) -> None:
    filing = wb["FilingInfo"]
    _blank_key_values(filing, range(8, filing.max_row + 1))
    signatory = wb["AuthorisedSignatory"]
    _blank_range(signatory, 4, 10, 3, signatory.max_column)
    report = wb["DNBS13"]
    _neutralise_report_header(report, "Reporting Period End Date :")
    _blank_range(report, 13, 14, 2, report.max_column)
    _blank_range(report, 16, report.max_row, 2, report.max_column)


def sanitize_dnbs4a(wb) -> None:
    filing = wb["FilingInfo"]
    _blank_key_values(filing, range(8, filing.max_row + 1))
    signatory = wb["AuthorisedSignatory"]
    _blank_range(signatory, 6, 13, 3, signatory.max_column)
    report = wb["DNBS4AShortTermDynamicLiquidity"]
    _neutralise_report_header(report, "Reporting Quarter End Date :")
    _blank_range(report, 11, 89, 3, report.max_column)
    _blank_range(report, 91, report.max_row, 3, report.max_column)


def sanitize_dnbs4b(wb) -> None:
    filing = wb["Filing Info"]
    _blank_key_values(filing, range(6, filing.max_row + 1))
    signatory = wb["AuthorisedSignatory"]
    _blank_range(signatory, 6, 13, 3, signatory.max_column)

    structural = wb["DNBS4BStructuralLiquidity"]
    _neutralise_report_header(structural, "Reporting Period End Date :")
    _blank_range(structural, 11, structural.max_row, 3, structural.max_column)

    irs = wb["DNBS4BIRS"]
    _neutralise_report_header(irs, "Reporting Period End Date :")
    _blank_range(irs, 11, 193, 3, irs.max_column)
    # Rows 194-197 contain the second table's title and column headers.
    _blank_range(irs, 198, irs.max_row, 3, irs.max_column)


SANITIZERS = {
    "DNBS13_Blank_Template.xlsx": sanitize_dnbs13,
    "DNBS4A_Blank_Template.xlsx": sanitize_dnbs4a,
    "DNBS4B_Blank_Template.xlsx": sanitize_dnbs4b,
}


def verify(path: Path) -> None:
    wb = openpyxl.load_workbook(path, data_only=False)
    problems: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                value = str(cell.value)
                if PAN_RE.search(value) or EMAIL_RE.search(value):
                    problems.append(f"{ws.title}!{cell.coordinate}={value!r}")
    if problems:
        raise RuntimeError(f"Sensitive filed values remain in {path.name}: {problems[:10]}")


def main(source_dir: Path) -> None:
    ASSET_DIR.mkdir(parents=True, exist_ok=True)
    for source_name, dest_name in FILES.items():
        source = source_dir / source_name
        if not source.exists():
            raise FileNotFoundError(source)
        wb = openpyxl.load_workbook(source, data_only=False)
        SANITIZERS[dest_name](wb)
        dest = ASSET_DIR / dest_name
        wb.save(dest)
        verify(dest)
        print(f"{source.name} -> {dest}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("Usage: build_regulatory_templates.py <extracted-directory>")
    main(Path(sys.argv[1]))
